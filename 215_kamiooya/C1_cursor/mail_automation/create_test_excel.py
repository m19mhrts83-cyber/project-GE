#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 新しいワークブックを作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "送信先一覧"

# ヘッダー行
headers = ["会社名", "担当者", "メールアドレス", "備考"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)

# テストデータ（自分のメールアドレスに置き換えてください）
test_data = [
    ["テスト1", "自分", "your.email@gmail.com", "テスト用"],
]

for row_idx, row_data in enumerate(test_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx).value = value

# 列幅を調整
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 20

# 保存
wb.save("テスト送信リスト.xlsx")
print("テスト用Excelファイルを作成しました: テスト送信リスト.xlsx")
