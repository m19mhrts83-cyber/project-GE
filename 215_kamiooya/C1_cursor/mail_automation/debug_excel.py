#!/usr/bin/env python3
"""Excel の G2 シートの中身を確認するデバッグ用スクリプト"""
import openpyxl
from pathlib import Path

# 送信スクリプトと同じパス
EXCEL = Path("/Users/matsunomasaharu/Library/CloudStorage/OneDrive-個人用/215_神・大家さん倶楽部/20_【空室対策】【修繕】【売却】/21_【空室対策】募集,ステージング,物件管理/★管理会社一覧.xlsx")

if not EXCEL.exists():
    print(f"ファイルが見つかりません: {EXCEL}")
    exit(1)

wb = openpyxl.load_workbook(EXCEL, data_only=True)
if "G2" not in wb.sheetnames:
    print(f"シート G2 がありません。利用可能: {wb.sheetnames}")
    exit(1)

sheet = wb["G2"]
print("=== 1行目（ヘッダー）===")
headers = [c.value for c in sheet[1]]
for i, h in enumerate(headers):
    print(f"  列{i}: {repr(h)}")

print("\n=== 2〜5行目（データ）===")
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
    print(f"  行{row_idx}: {row[:10]}")  # 最初の10列のみ

print("\n=== メール関連列の値（2〜10行目）===")
for idx, h in enumerate(headers):
    if h and ('メール' in str(h) or 'mail' in str(h).lower() or 'email' in str(h).lower()):
        print(f"  列 '{h}' (idx={idx}):")
        for r in range(2, 11):
            val = sheet.cell(row=r, column=idx + 1).value
            has_at = '@' in str(val) if val else False
            print(f"    行{r}: {repr(val)} {' ← @あり' if has_at else ''}")
