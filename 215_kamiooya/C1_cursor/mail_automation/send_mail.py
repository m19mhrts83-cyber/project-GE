#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
メール自動送信スクリプト

MDファイルから件名と本文を読み取り、
Excelファイルからメールアドレス・会社名・担当者名を取得して、
Gmail APIで1社ずつ個別にメールを送信するスクリプト

各メールの本文冒頭に「会社名　担当者名　様」を自動挿入します。

--exclude "会社名,担当者名" で、送信しない宛先を指定可能（複数可）。
チャットで「この人には送らないで」と指示されたときに利用する。
"""

import os
import sys
import argparse
import json
import pickle
import time
from pathlib import Path
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
import base64

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

try:
    import openpyxl
except ImportError:
    print("エラー: openpyxlがインストールされていません。")
    print("以下のコマンドでインストールしてください:")
    print("  pip install openpyxl")
    sys.exit(1)

# Gmail APIのスコープ
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

# スクリプトのディレクトリパス
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / 'logs'

# 認証ファイル（現在の設定: パートナー・いけともと同じ。3日ごとのトークン自動更新の対象）
DEFAULT_CREDENTIALS_DIR = SCRIPT_DIR.parent / '1b_Cursorマニュアル'
CREDENTIALS_FILE = Path(os.environ.get('GMAIL_CREDENTIALS_PATH', str(DEFAULT_CREDENTIALS_DIR / 'credentials.json')))
TOKEN_FILE = Path(os.environ.get('GMAIL_TOKEN_PATH', str(DEFAULT_CREDENTIALS_DIR / 'token.json')))

# --- 以下は旧: 空室対策専用（mail_automation 内 credentials.json / token.pickle）
# (a) 現在は使用していない
# (b) 認証は 1b_Cursorマニュアルの credentials.json / token.json に統一された
CREDENTIALS_FILE_LEGACY = SCRIPT_DIR / 'credentials.json'
TOKEN_FILE_LEGACY = SCRIPT_DIR / 'token.pickle'


def authenticate_gmail():
    """Gmail APIの認証を行う（1b_Cursorマニュアルの token.json を使用。未設定時は旧 token.pickle にフォールバック）"""
    credentials_path = CREDENTIALS_FILE
    token_path = TOKEN_FILE

    # 現在の設定: token.json（1b_Cursorマニュアル）で認証
    creds = None
    if token_path.exists() and token_path.suffix == '.json':
        try:
            token_data = json.loads(token_path.read_text(encoding='utf-8'))
            creds_data = dict(token_data)
            if 'client_id' not in creds_data and credentials_path.exists():
                cred_data = json.loads(credentials_path.read_text(encoding='utf-8'))
                client = cred_data.get('installed') or cred_data.get('web', {})
                creds_data['client_id'] = client.get('client_id')
                creds_data['client_secret'] = client.get('client_secret')
                creds_data['token_uri'] = 'https://oauth2.googleapis.com/token'
                if 'access_token' in creds_data and 'token' not in creds_data:
                    creds_data['token'] = creds_data['access_token']
            creds = Credentials.from_authorized_user_info(creds_data, SCOPES)
        except Exception:
            creds = None

    if creds and creds.valid:
        return creds
    if creds and creds.expired and creds.refresh_token:
        print("認証情報を更新しています...")
        creds.refresh(Request())
        if token_path.suffix == '.json':
            with open(token_path, 'w', encoding='utf-8') as f:
                f.write(creds.to_json())
        print("認証が完了しました。")
        return creds
    if credentials_path.exists():
        print("初回認証を開始します...")
        print("ブラウザが開きますので、Googleアカウントで認証してください。")
        flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
        creds = flow.run_local_server(port=0)
        if token_path.suffix == '.json':
            with open(token_path, 'w', encoding='utf-8') as f:
                f.write(creds.to_json())
        print("認証が完了しました。")
        return creds

    # フォールバック: 旧 空室対策専用 token.pickle
    # (a) 現在は使用していない（通常は上記 token.json を使用）
    # (b) 別のものに変更された（1b_Cursorマニュアルの credentials.json / token.json に統一）
    return _authenticate_gmail_legacy()


def _authenticate_gmail_legacy():
    """旧: mail_automation 内 credentials.json / token.pickle で認証。（現在は使用していない。1b_Cursorマニュアルの token.json に変更済み）"""
    creds = None
    if TOKEN_FILE_LEGACY.exists():
        with open(TOKEN_FILE_LEGACY, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("認証情報を更新しています...")
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE_LEGACY.exists():
                print(f"エラー: credentials.jsonが見つかりません。")
                print(f"期待されるパス（現設定）: {CREDENTIALS_FILE}")
                print(f"旧パス: {CREDENTIALS_FILE_LEGACY}")
                print("\n1b_Cursorマニュアルに credentials.json と token.json を用意するか、")
                print("README_Gmail_API_Setup.mdを参照して認証情報を作成してください。")
                sys.exit(1)
            print("初回認証を開始します（旧 token.pickle 用）...")
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE_LEGACY), SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE_LEGACY, 'wb') as token:
            pickle.dump(creds, token)
        print("認証が完了しました。")
    return creds


def read_markdown_file(md_file_path):
    """
    MDファイルを読み込み、件名と本文を抽出する
    
    Args:
        md_file_path: MDファイルのパス
    
    Returns:
        tuple: (件名, 本文)
    """
    md_path = Path(md_file_path)
    
    if not md_path.exists():
        raise FileNotFoundError(f"MDファイルが見つかりません: {md_file_path}")
    
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    if not lines:
        raise ValueError("MDファイルが空です")
    
    # 1行目を件名、2行目以降を本文とする
    subject = lines[0].strip()
    body = ''.join(lines[1:]).strip()
    
    return subject, body


def read_recipients_from_excel(excel_file_path, sheet_name=None, email_column=None,
                               company_column=None, contact_column=None):
    """
    Excelファイルからメールアドレス・会社名・担当者名を読み取る
    
    Args:
        excel_file_path: Excelファイルのパス
        sheet_name: シート名（省略時は最初のシート）
        email_column: メールアドレス列の名前（省略時は自動検出）
        company_column: 会社名列の名前（省略時は自動検出）
        contact_column: 担当者名列の名前（省略時は自動検出）
    
    Returns:
        list: 宛先情報の辞書リスト [{'email': ..., 'company': ..., 'contact': ...}, ...]
    """
    excel_path = Path(excel_file_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_file_path}")
    
    # Excelファイルを読み込む
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    
    # シートを選択
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が見つかりません。\n"
                           f"利用可能なシート: {', '.join(workbook.sheetnames)}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
        print(f"シート '{sheet.title}' を使用します")
    
    # ヘッダー行を取得（1行目）
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # --- メールアドレス列を特定 ---
    # 「個別メール」より「mail」を優先（G2シート等で宛先が mail 列に入っているため）
    email_col_idx = None
    if email_column:
        try:
            email_col_idx = headers.index(email_column)
        except ValueError:
            raise ValueError(f"列 '{email_column}' が見つかりません。\n"
                           f"利用可能な列: {', '.join([str(h) for h in headers if h])}")
    else:
        # 優先: 列名が "mail" / "メールアドレス" / "email" のもの（完全一致）
        preferred = ("mail", "メールアドレス", "email")
        for idx, header in enumerate(headers):
            if not header:
                continue
            h = str(header).strip().lower().replace("\n", "")
            if h in preferred:
                email_col_idx = idx
                print(f"メールアドレス列として '{header}' を使用します")
                break
        # 見つからなければ従来どおり「メール」「mail」「email」を含む最初の列
        if email_col_idx is None:
            for idx, header in enumerate(headers):
                if header and ('メール' in str(header).lower() or
                              'email' in str(header).lower() or
                              'mail' in str(header).lower()):
                    email_col_idx = idx
                    print(f"メールアドレス列として '{header}' を使用します")
                    break
    
    if email_col_idx is None:
        raise ValueError("メールアドレスの列が見つかりません。\n"
                        "--email-column オプションで列名を指定してください。")
    
    # --- 会社名列を特定 ---
    company_col_idx = None
    if company_column:
        try:
            company_col_idx = headers.index(company_column)
        except ValueError:
            print(f"⚠️  会社名列 '{company_column}' が見つかりません。会社名なしで送信します。")
    else:
        for idx, header in enumerate(headers):
            if header and '会社' in str(header):
                company_col_idx = idx
                print(f"会社名列として '{header}' を使用します")
                break
    
    if company_col_idx is None:
        print("⚠️  会社名列が見つかりません。会社名なしで送信します。")
    
    # --- 担当者名列を特定 ---
    contact_col_idx = None
    if contact_column:
        try:
            contact_col_idx = headers.index(contact_column)
        except ValueError:
            print(f"⚠️  担当者名列 '{contact_column}' が見つかりません。担当者名なしで送信します。")
    else:
        for idx, header in enumerate(headers):
            if header and '担当' in str(header):
                contact_col_idx = idx
                print(f"担当者名列として '{header}' を使用します")
                break
    
    if contact_col_idx is None:
        print("⚠️  担当者名列が見つかりません。担当者名なしで送信します。")
    
    # 宛先情報を抽出（2行目以降）
    recipients = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email = row[email_col_idx] if email_col_idx is not None and email_col_idx < len(row) else None
        if email and '@' in str(email):
            company = ''
            contact = ''
            if company_col_idx is not None and company_col_idx < len(row) and row[company_col_idx]:
                company = str(row[company_col_idx]).strip()
            if contact_col_idx is not None and contact_col_idx < len(row) and row[contact_col_idx]:
                contact = str(row[contact_col_idx]).strip()
            
            recipients.append({
                'email': str(email).strip(),
                'company': company,
                'contact': contact,
            })
    
    if not recipients:
        raise ValueError("メールアドレスが1件も見つかりませんでした")
    
    return recipients


def filter_excluded_recipients(recipients, exclude_list):
    """
    除外指定（会社名・担当者名のペア）に一致する宛先を送信対象から外す。

    Args:
        recipients: read_recipients_from_excel の戻り値
        exclude_list: 除外ペアのリスト。各要素は (会社名パターン, 担当者名パターン)。
                     空文字のパターンは「任意」とみなす。部分一致・大文字小文字区別なし。

    Returns:
        tuple: (フィルタ後の宛先リスト, 除外された宛先のリスト)
    """
    def normalize(s):
        return (s or "").strip()

    def matches(recipient, company_pattern, contact_pattern):
        r_company = normalize(recipient.get("company") or "")
        r_contact = normalize(recipient.get("contact") or "")
        c_ok = not company_pattern or (company_pattern.lower() in r_company.lower())
        t_ok = not contact_pattern or (contact_pattern.lower() in r_contact.lower())
        return c_ok and t_ok

    kept = []
    excluded = []
    for r in recipients:
        skip = False
        for (company_pattern, contact_pattern) in exclude_list:
            if matches(r, company_pattern, contact_pattern):
                skip = True
                excluded.append(r)
                break
        if not skip:
            kept.append(r)
    return kept, excluded


def build_personalized_body(body, company, contact):
    """
    本文の先頭に会社名・担当者名の宛名を挿入する
    
    Args:
        body: 元の本文
        company: 会社名
        contact: 担当者名
    
    Returns:
        str: 宛名付きの本文
    """
    # 宛名行を組み立てる（会社名と担当者名を別行に）
    if company and contact:
        greeting = f"{company}\n{contact} 様\n\n"
    elif company:
        greeting = f"{company}\nご担当者 様\n\n"
    elif contact:
        greeting = f"{contact} 様\n\n"
    else:
        greeting = ""
    
    return greeting + body


def create_message(subject, body, to_address):
    """
    個別送信用のメールメッセージを作成する
    
    Args:
        subject: 件名
        body: 本文（宛名挿入済み）
        to_address: 送信先メールアドレス（1件）
    
    Returns:
        dict: Gmail APIに送信するメッセージ
    """
    message = MIMEMultipart()
    
    if not to_address:
        raise ValueError("送信先が指定されていません")
    
    message['To'] = to_address
    message['Subject'] = subject
    
    # 本文を追加
    message.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # Base64エンコード
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
    
    return {'raw': raw_message}


def send_email(service, message):
    """
    Gmail APIでメールを送信する
    
    Args:
        service: Gmail APIサービスオブジェクト
        message: 送信するメッセージ
    
    Returns:
        dict: 送信結果
    """
    try:
        sent_message = service.users().messages().send(
            userId='me', body=message).execute()
        return sent_message
    except HttpError as error:
        print(f'エラーが発生しました: {error}')
        raise


def save_log(subject, num_recipients, status, error_msg=None, scheduled_time=None):
    """
    送信ログを保存する
    
    Args:
        subject: 件名
        num_recipients: 送信先件数
        status: 送信状態（success/failed/scheduled）
        error_msg: エラーメッセージ（失敗時）
        scheduled_time: スケジュール送信時刻
    """
    LOG_DIR.mkdir(exist_ok=True)
    log_file = LOG_DIR / 'send_history.log'
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_entry = f"[{timestamp}] {status.upper()} - 件名: {subject} - 送信先: {num_recipients}件"
    
    if scheduled_time:
        log_entry += f" - スケジュール: {scheduled_time}"
    
    if error_msg:
        log_entry += f" - エラー: {error_msg}"
    
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(log_entry + '\n')


def display_preview(subject, body, recipients, full_preview=False):
    """
    送信内容のプレビューを表示する
    
    Args:
        subject: 件名
        body: 元の本文（宛名挿入前）
        recipients: 宛先情報の辞書リスト
        full_preview: True の場合、本文を全文表示
    """
    print("\n" + "=" * 70)
    print("📧 送信内容のプレビュー（個別送信モード）")
    print("=" * 70)
    print(f"\n【件名】\n{subject}")
    print("\n" + "-" * 70)
    
    # 1件目のサンプルで宛名付き本文を表示
    sample = recipients[0]
    sample_body = build_personalized_body(body, sample['company'], sample['contact'])
    
    if full_preview:
        print(f"\n【本文サンプル（1件目）】（全{len(sample_body)}文字）")
        print(sample_body)
    else:
        print(f"\n【本文サンプル（1件目）】（{len(sample_body)}文字、先頭400文字を表示）")
        preview_text = sample_body[:400]
        print(preview_text)
        if len(sample_body) > 400:
            print("\n... (省略されました。全文を見るには --full-preview を使用)")
    
    print("\n" + "-" * 70)
    print(f"\n【送信先】個別送信: {len(recipients)}件")
    
    # 最初の10件を表示
    display_count = min(10, len(recipients))
    for i, r in enumerate(recipients[:display_count], 1):
        company = r['company'] or '(会社名なし)'
        contact = r['contact'] or '(担当者なし)'
        print(f"  {i:2d}. {r['email']}  ← {company} / {contact}")
    
    if len(recipients) > display_count:
        print(f"  ... 他 {len(recipients) - display_count}件")
    
    print("\n" + "=" * 70)


def parse_schedule_time(schedule_str):
    """
    スケジュール時刻の文字列をパースする
    
    Args:
        schedule_str: 日時文字列（例: "2026-01-30 14:30", "14:30", "30分後"）
    
    Returns:
        datetime: パースされた日時
    """
    now = datetime.now()
    
    # "XX分後" または "XX時間後" の形式
    if "分後" in schedule_str:
        minutes = int(schedule_str.replace("分後", "").strip())
        return now + timedelta(minutes=minutes)
    elif "時間後" in schedule_str:
        hours = int(schedule_str.replace("時間後", "").strip())
        return now + timedelta(hours=hours)
    
    # "HH:MM" 形式（今日または明日）
    if ":" in schedule_str and "-" not in schedule_str:
        time_str = schedule_str.strip()
        hour, minute = map(int, time_str.split(":"))
        scheduled = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
        
        # 過去の時刻の場合は翌日にする
        if scheduled < now:
            scheduled += timedelta(days=1)
        
        return scheduled
    
    # "YYYY-MM-DD HH:MM" 形式
    try:
        return datetime.strptime(schedule_str, "%Y-%m-%d %H:%M")
    except ValueError:
        pass
    
    # "MM/DD HH:MM" 形式
    try:
        parsed = datetime.strptime(schedule_str, "%m/%d %H:%M")
        scheduled = parsed.replace(year=now.year)
        
        # 過去の日付の場合は翌年にする
        if scheduled < now:
            scheduled = scheduled.replace(year=now.year + 1)
        
        return scheduled
    except ValueError:
        pass
    
    raise ValueError(f"日時のフォーマットが正しくありません: {schedule_str}\n"
                    f"使用可能な形式:\n"
                    f"  - YYYY-MM-DD HH:MM (例: 2026-01-30 14:30)\n"
                    f"  - MM/DD HH:MM (例: 01/30 14:30)\n"
                    f"  - HH:MM (例: 14:30)\n"
                    f"  - XX分後 (例: 30分後)\n"
                    f"  - XX時間後 (例: 2時間後)")


def wait_until_scheduled_time(scheduled_time):
    """
    指定された時刻まで待機する
    
    Args:
        scheduled_time: 送信予定時刻
    """
    now = datetime.now()
    wait_seconds = (scheduled_time - now).total_seconds()
    
    if wait_seconds <= 0:
        print("⚠️  指定された時刻は既に過ぎています。すぐに送信します。")
        return
    
    print(f"\n⏰ スケジュール送信: {scheduled_time.strftime('%Y年%m月%d日 %H:%M')}")
    print(f"   待機時間: {int(wait_seconds // 60)}分{int(wait_seconds % 60)}秒")
    print("\n   待機中... (Ctrl+C でキャンセル)")
    
    try:
        # 進捗表示付きで待機
        interval = 30  # 30秒ごとに更新
        elapsed = 0
        
        while elapsed < wait_seconds:
            remaining = wait_seconds - elapsed
            remaining_minutes = int(remaining // 60)
            remaining_seconds = int(remaining % 60)
            
            print(f"\r   残り時間: {remaining_minutes:3d}分 {remaining_seconds:02d}秒", end="", flush=True)
            
            sleep_time = min(interval, remaining)
            time.sleep(sleep_time)
            elapsed += sleep_time
        
        print("\r   " + " " * 50, end="\r")  # クリア
        print("✓ スケジュール時刻になりました。送信を開始します。\n")
        
    except KeyboardInterrupt:
        print("\n\n送信をキャンセルしました")
        sys.exit(0)


def main():
    parser = argparse.ArgumentParser(
        description='MDファイルとExcelファイルを使って、1社ずつ個別にメールを送信します',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  # 基本的な使い方（会社名・担当者名を自動検出して個別送信）
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx
  
  # シート・列を明示指定
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx --sheet-name "G2" --email-column "mail" --company-column "会社名" --contact-column "担当"
  
  # 本文全文をプレビュー
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx --full-preview
  
  # 送信時刻を指定
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx --schedule "14:30"
  
  # プレビューのみ（送信しない）
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx --dry-run
  
  # 送信間隔を変更（デフォルト2秒）
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx --interval 3
  
  # 特定の会社・担当者を除外して送信（チャットで「この人には送らないで」と指示するとき）
  python send_mail.py --md-file メール.md --excel-file 送信先.xlsx --exclude "株式会社A,山田" --exclude "B不動産,鈴木"
        """
    )
    
    parser.add_argument('--md-file', required=True,
                       help='メール内容が書かれたMDファイルのパス（1行目=件名、2行目以降=本文）')
    parser.add_argument('--excel-file', required=True,
                       help='メールアドレス一覧が含まれるExcelファイルのパス')
    parser.add_argument('--sheet-name', default=None,
                       help='Excelのシート名（省略時は最初のシート）')
    parser.add_argument('--email-column', default=None,
                       help='メールアドレス列の名前（省略時は自動検出）')
    parser.add_argument('--company-column', default=None,
                       help='会社名列の名前（省略時は自動検出）')
    parser.add_argument('--contact-column', default=None,
                       help='担当者名列の名前（省略時は自動検出）')
    parser.add_argument('--dry-run', action='store_true',
                       help='実際には送信せず、送信内容のプレビューのみ表示')
    parser.add_argument('--full-preview', action='store_true',
                       help='本文を全文表示する（デフォルトは先頭300文字のみ）')
    parser.add_argument('--schedule', '--send-at', dest='schedule',
                       help='送信時刻を指定（例: "14:30", "2026-01-30 14:30", "30分後", "2時間後"）')
    parser.add_argument('--yes', '-y', action='store_true',
                       help='確認なしで自動的に送信する')
    parser.add_argument('--interval', type=float, default=2.0,
                       help='メール送信間隔（秒）。Gmail レート制限対策（デフォルト: 2秒）')
    parser.add_argument('--exclude', action='append', default=[], metavar='会社名,担当者名',
                       help='送信しない宛先を指定（複数可）。会社名と担当者名をカンマ区切りで指定。'
                            '部分一致で判定。例: --exclude "株式会社A,山田" --exclude "B不動産,鈴木"')

    args = parser.parse_args()
    
    try:
        # MDファイルから件名と本文を読み込む
        print("=" * 70)
        print("📄 MDファイルを読み込んでいます...")
        subject, body = read_markdown_file(args.md_file)
        print(f"✓ 件名: {subject}")
        print(f"✓ 本文: {len(body)}文字")
        
        # Excelファイルから宛先情報を読み込む
        print("\n" + "=" * 70)
        print("📊 Excelファイルから宛先情報を読み込んでいます...")
        recipients = read_recipients_from_excel(
            args.excel_file, args.sheet_name, args.email_column,
            args.company_column, args.contact_column
        )
        print(f"✓ {len(recipients)}件の宛先を取得しました")

        # 除外指定の適用（会社名・担当者名で送信しない人を除外）
        exclude_list = []
        for s in (args.exclude or []):
            parts = [p.strip() for p in str(s).split(",", 1)]
            company_pattern = parts[0] if len(parts) > 0 else ""
            contact_pattern = parts[1] if len(parts) > 1 else ""
            exclude_list.append((company_pattern, contact_pattern))
        if exclude_list:
            recipients, excluded = filter_excluded_recipients(recipients, exclude_list)
            if excluded:
                print(f"✓ 除外指定により {len(excluded)}件を送信対象から外しました:")
                for r in excluded[:10]:
                    c = (r.get("company") or "(会社名なし)")
                    t = (r.get("contact") or "(担当者なし)")
                    print(f"    - {c} / {t}  ({r.get('email')})")
                if len(excluded) > 10:
                    print(f"    ... 他 {len(excluded) - 10}件")
            if not recipients:
                print("エラー: 除外後に送信先が0件になりました。")
                sys.exit(1)

        # プレビュー表示
        display_preview(subject, body, recipients, args.full_preview)
        
        # Dry-runモードの場合はここで終了
        if args.dry_run:
            print("\n[DRY-RUN モード] 実際の送信は行いませんでした")
            return
        
        # スケジュール送信の場合の処理
        scheduled_time = None
        if args.schedule:
            try:
                scheduled_time = parse_schedule_time(args.schedule)
                print(f"\n📅 送信予定時刻: {scheduled_time.strftime('%Y年%m月%d日 %H時%M分')}")
            except ValueError as e:
                print(f"\nエラー: {e}")
                sys.exit(1)
        
        # 送信確認
        print("\n" + "=" * 70)
        if scheduled_time:
            print(f"上記の内容を {scheduled_time.strftime('%Y年%m月%d日 %H:%M')} に {len(recipients)}社へ個別送信します。")
        else:
            print(f"上記の内容を今すぐ {len(recipients)}社へ個別送信します。")
        print(f"（送信間隔: {args.interval}秒）")
        
        # --yes オプションが指定されている場合は確認をスキップ
        if args.yes:
            print("✓ 自動承認モード: 送信を開始します")
        else:
            confirmation = input("送信してもよろしいですか? (yes/no): ").strip().lower()
            
            if confirmation not in ['yes', 'y']:
                print("送信をキャンセルしました")
                return
        
        # スケジュール送信の場合は待機
        if scheduled_time:
            wait_until_scheduled_time(scheduled_time)
        
        # Gmail APIの認証
        print("\n" + "=" * 70)
        print("🔐 Gmail APIに接続しています...")
        creds = authenticate_gmail()
        service = build('gmail', 'v1', credentials=creds)
        print("✓ 接続しました")
        
        # 1社ずつ個別にメールを送信
        print(f"\n📤 メールを個別送信しています（全{len(recipients)}件）...")
        print("-" * 70)
        
        success_count = 0
        fail_count = 0
        failed_recipients = []
        
        for i, recipient in enumerate(recipients, 1):
            company = recipient['company'] or '(会社名なし)'
            contact = recipient['contact'] or '(担当者なし)'
            email = recipient['email']
            
            try:
                # 宛名付きの本文を作成
                personalized_body = build_personalized_body(
                    body, recipient['company'], recipient['contact']
                )
                
                # メッセージを作成
                message = create_message(
                    subject=subject,
                    body=personalized_body,
                    to_address=email
                )
                
                # メール送信
                result = send_email(service, message)
                success_count += 1
                print(f"  ✅ [{i}/{len(recipients)}] {company} / {contact} → {email}")
                
            except Exception as e:
                fail_count += 1
                failed_recipients.append({'recipient': recipient, 'error': str(e)})
                print(f"  ❌ [{i}/{len(recipients)}] {company} / {contact} → {email}  エラー: {e}")
            
            # レート制限対策: 最後の1件以外は待機
            if i < len(recipients):
                time.sleep(args.interval)
        
        # 送信結果サマリー
        print("\n" + "=" * 70)
        print("📊 送信結果サマリー")
        print("=" * 70)
        print(f"  ✅ 成功: {success_count}件")
        if fail_count > 0:
            print(f"  ❌ 失敗: {fail_count}件")
            for f in failed_recipients:
                r = f['recipient']
                print(f"     - {r['company'] or '?'} / {r['contact'] or '?'} ({r['email']}): {f['error']}")
        
        if scheduled_time:
            print(f"  ⏰ 送信時刻: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}")
        
        if fail_count == 0:
            print("\n✅ 全件送信が完了しました！")
        else:
            print(f"\n⚠️  {fail_count}件の送信に失敗しました。上記のエラーを確認してください。")
        print("=" * 70)
        
        # ログ保存
        status = 'success' if fail_count == 0 else 'partial'
        error_msg = f"{fail_count}件失敗" if fail_count > 0 else None
        save_log(subject, success_count, status, error_msg=error_msg,
                scheduled_time=scheduled_time.strftime('%Y-%m-%d %H:%M') if scheduled_time else None)
        
    except FileNotFoundError as e:
        print(f"\n❌ エラー: {e}")
        save_log('N/A', 0, 'failed', error_msg=str(e))
        sys.exit(1)
    except ValueError as e:
        print(f"\n❌ エラー: {e}")
        save_log('N/A', 0, 'failed', error_msg=str(e))
        sys.exit(1)
    except HttpError as e:
        print(f"\n❌ Gmail API エラー: {e}")
        save_log(subject if 'subject' in locals() else 'N/A', 
                len(recipients) if 'recipients' in locals() else 0,
                'failed', error_msg=str(e))
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\n⚠️  ユーザーによってキャンセルされました")
        if 'success_count' in locals() and success_count > 0:
            print(f"   ※ {success_count}件は送信済みです")
            save_log(subject, success_count, 'interrupted',
                    error_msg=f"ユーザーキャンセル（{success_count}件送信済み）")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 予期しないエラー: {e}")
        import traceback
        traceback.print_exc()
        save_log('N/A', 0, 'failed', error_msg=str(e))
        sys.exit(1)


if __name__ == '__main__':
    main()
