#!/usr/bin/env python3
"""銀行提出用：法人履歴書・事業計画書テンプレへ記入案を反映（openpyxl）。"""
from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path

import openpyxl

BASE = Path(
    "/Users/matsunomasaharu2/Library/CloudStorage/OneDrive-個人用/215_神・大家さん倶楽部"
)
DIR_70 = BASE / "70_経歴書,事業経歴書"
TEMPL_H = DIR_70 / "【銀行が喜ぶ経歴書】法人履歴書テンプレ.xlsx"
TEMPL_J = DIR_70 / "【銀行が喜ぶ事業計画書】事業計画書テンプレ.xlsx"
OUT_H = DIR_70 / "法人履歴書_リビングサポート松_202606.xlsx"
OUT_J = DIR_70 / "事業計画書_リビングサポート松_202606.xlsx"

# 第1期決算（knees bee 申告書一式と整合）
P1 = {
    "sales": 973_104,
    "sales_k": 973,
    "operating_profit": 99_419,
    "operating_k": 99,
    "ordinary_profit": -313_732,
    "ordinary_k": -314,
    "net_income": -337_332,
    "net_k": -337,
    "interest": 281_274,
    "capital": 100_000,
    "orix_loan": 61_410_383,
    "officer_loan": 8_325_103,
    "total_loan": 69_735_486,
    "deposit": 1_204_000,  # glandole1 現預金1204千円
}

FORECAST = [
    {"label": "第1期（実績）", "period": "2025/1～2025/5", "sales_k": 973, "op_k": 99, "net_k": -337},
    {"label": "第2期（見込）", "period": "2025/6～2026/5", "sales_k": 3886, "op_k": 200, "net_k": -100},
    {"label": "第3期（見込）", "period": "2026/6～2027/5", "sales_k": 4000, "op_k": 350, "net_k": 50},
    {"label": "第4期（見込）", "period": "2027/6～2028/5", "sales_k": 4200, "op_k": 500, "net_k": 200},
    {"label": "第5期（見込）", "period": "2028/6～2029/5", "sales_k": 4272, "op_k": 550, "net_k": 250},
]

COMPANY = {
    "name": "株式会社リビングサポート松",
    "rep": "代表取締役　松野　千景",
    "established": datetime(2025, 1, 6),
    "address": "愛知県豊明市間米町唐竹３４２番地２５",
    "settlement": "5月末",
    "period_label": "現在2期",
    "capital": "10万円",
    "shareholder": "代表取締役　松野　千景以外の株主は松野　真治（100%）",
}

PROPERTY = {
    "name": "Grandole志賀本通Ⅰ",
    "location": "名古屋市守山区志賀本通",
    "type": "賃貸アパート",
    "structure": "木造",
    "acquired": "令和7年2月28日",
}

BUSINESS_DESC = (
    "・不動産の所有・賃貸経営。Grandole志賀本通Ⅰを法人所有し、賃貸管理を行っています。\n"
    "・不動産の管理業務（当社所有物件および将来取得物件）。\n"
    "・収益不動産の段階的な取得・安定運営により、継続的な賃料収入を確保します。"
)

VISION = (
    "・所有不動産の安定経営を目標とし、満室経営・黒字継続を実現する。\n"
    "・愛知県内を中心に収益不動産を段階的に拡大し、返済原資を確保する。\n"
    "・取引金融機関との信頼関係を構築し、計画的な資金調達・返済を行う。"
)

FUTURE_POLICY = (
    "愛知県内を中心に収益不動産を段階的に拡大し、満室・黒字継続を目指します。"
    "返済原資は物件キャッシュフローに加え、株主（松野真治）の不動産所得および計画的な自己資金蓄積で確保します。"
    "法人所有物件の安定稼働を最優先とし、取得・売却は事業目的と税務上の合理性を踏まえ計画的に行います。"
    "取引金融機関との継続的な情報共有を通じ、返済計画の透明性を高めます。"
)

OFFICER_LOAN_NOTE = (
    "決算書上、上記以外の借入金は、役員（株主・松野真治）より貸し付けたもの（8,325,103円）。"
    "利率・返済方法・期限は別途契約書に基づく（銀行提出前に記入要）。"
)

RENTAL_OPS = (
    "【空室対策】入居促進・条件見直し・管理会社（ミニテック・Tcell・LEAF）との連携により稼働率改善を図る。\n"
    "【修繕】計画的な修繕・原状回復を管理会社と連携して実施。\n"
    "【経験】個人事業（松愛不動産・令和2年開業）および個人所有物件の賃貸経営実績を活かす。"
)


def fill_hojin_rirekisho(ws) -> None:
    ws["H11"] = COMPANY["name"]
    ws["E29"] = COMPANY["name"]
    ws["E30"] = COMPANY["rep"]
    ws["E31"] = COMPANY["established"]
    ws["E32"] = COMPANY["address"]
    ws["E33"] = f"毎年{COMPANY['settlement']}"
    ws["E34"] = COMPANY["period_label"]
    ws["E35"] = COMPANY["capital"]
    ws["E36"] = BUSINESS_DESC
    ws["E37"] = VISION
    ws["E39"] = "一期（短縮：令和7年1月6日〜5月31日）"
    ws["E40"] = f"{P1['sales']:,}円"
    ws["E41"] = f"{P1['operating_profit']:,}円"
    ws["E42"] = f"{P1['ordinary_profit']:,}円"
    ws["E43"] = f"{P1['net_income']:,}円（赤字）"

    ws["A48"] = "不動産賃貸・管理"
    ws["C48"] = (
        "Grandole志賀本通Ⅰ（賃貸アパート）の賃貸経営および管理業務。"
        "売上は主に賃料収入（第1期973,104円）。"
    )
    ws["A52"] = "売上構成：不動産賃貸収入100%（第1期）"

    ws["E57"] = f"{P1['deposit'] // 10000}万円"
    ws["E58"] = f"{P1['deposit'] // 10000}万円"
    ws["C60"] = PROPERTY["type"]
    ws["E60"] = PROPERTY["name"]
    ws["G60"] = PROPERTY["location"]
    ws["E63"] = f"・{PROPERTY['name']}は、オリックス銀行借入（{P1['orix_loan']:,}円）の担保物件。"

    ws["C68"] = "オリックス銀行"
    ws["E68"] = PROPERTY["name"]
    ws["H68"] = P1["orix_loan"]
    ws["C69"] = "役員借入（松野真治）"
    ws["E69"] = "-"
    ws["H69"] = P1["officer_loan"]
    ws["C70"] = None
    ws["E70"] = None
    ws["H70"] = None
    ws["C71"] = None
    ws["E71"] = None
    ws["H71"] = None
    ws["H74"] = P1["total_loan"]
    ws["H76"] = OFFICER_LOAN_NOTE


def fill_jigyo_keikaku(ws) -> None:
    ws["I29"] = COMPANY["name"]
    ws["I30"] = COMPANY["rep"]
    ws["I31"] = COMPANY["established"]
    ws["I32"] = COMPANY["address"]
    ws["I33"] = COMPANY["settlement"]
    ws["I34"] = COMPANY["period_label"]
    ws["I35"] = COMPANY["capital"]
    ws["I36"] = (
        "事業１　不動産賃貸・管理業\n"
        "　・Grandole志賀本通Ⅰの賃貸経営および管理業務を行っています。\n\n"
        "事業２・３　該当事業なし"
    )
    ws["I37"] = (
        "株式会社リビングサポート松は、不動産の賃貸・管理を通じて"
        "地域の住環境維持に貢献します。SDGs「11：住み続けられるまちづくり」に寄与します。"
    )
    ws["I38"] = VISION
    ws["I41"] = "一期"
    ws["I42"] = "2025/1～2025/5（短縮第1期）"
    ws["I43"] = P1["sales_k"]
    ws["I44"] = 0
    ws["I45"] = 0
    ws["I46"] = P1["sales_k"]
    ws["I48"] = P1["operating_k"]
    ws["I49"] = 0
    ws["I50"] = 0
    ws["I51"] = P1["operating_k"]
    ws["I54"] = P1["ordinary_k"]
    ws["I55"] = P1["net_k"]
    ws["I56"] = "短縮第1期の立上げ期赤字。第2期以降は稼働率改善を見込む。"
    ws["I59"] = "株主　松野　真治　100%（代表取締役は松野　千景）"

    ws["C88"] = (
        "【今期（第2期）】\n\n"
        "目標：Grandole志賀本通Ⅰの稼働率改善により、年間売上高約3,885,600円（NET家賃・事業計画ベース）を目指す。\n\n"
        "1  空室対策・入居促進により稼働率77%前後を目標とする。\n"
        "2  オリックス銀行借入の返済を継続し、返済原資は物件CF＋個人不動産所得で確保。\n"
        "3  愛知県内を中心に、計画的な収益不動産の拡大を検討（段階的・慎重に実施）。"
    )

    ws["E738"] = "不動産賃貸・管理"
    ws["E739"] = "賃貸用不動産の安定供給および入居者の快適な住環境の維持"
    ws["E740"] = (
        "Grandole志賀本通Ⅰの賃貸経営、管理会社との連携による入居者対応・"
        "家賃回収・修繕管理を行う事業。"
    )
    ws["E746"] = P1["sales_k"]
    ws["H746"] = "Grandole志賀本通Ⅰ 賃料収入（短縮第1期実績）"
    ws["E747"] = 0
    ws["H747"] = "管理収入（第1期は賃貸収入に含む）"

    ws["A752"] = RENTAL_OPS.split("\n")[0].replace("【空室対策】", "")
    ws["A753"] = RENTAL_OPS.split("\n")[1].replace("【修繕】", "")
    ws["A754"] = RENTAL_OPS.split("\n")[2].replace("【経験】", "")

    ws["E763"] = "PayPay銀行（法人口座）"
    ws["I763"] = f"{P1['deposit'] // 10000}万円"
    ws["I764"] = f"{P1['deposit'] // 10000}万円"

    ws["C770"] = PROPERTY["type"]
    ws["E770"] = PROPERTY["name"]
    ws["G770"] = PROPERTY["location"]
    ws["J770"] = PROPERTY["structure"]
    ws["I773"] = f"※{PROPERTY['name']}はオリックス銀行借入の担保物件"

    ws["C777"] = "オリックス銀行"
    ws["E777"] = PROPERTY["name"]
    ws["H777"] = P1["orix_loan"]
    ws["C778"] = "役員借入（松野真治）"
    ws["E778"] = "-"
    ws["H778"] = P1["officer_loan"]
    ws["C779"] = None
    ws["E779"] = None
    ws["H779"] = None
    ws["C780"] = None
    ws["E780"] = None
    ws["H780"] = None
    ws["C781"] = None
    ws["E781"] = None
    ws["H781"] = None
    ws["H782"] = P1["total_loan"]

    ws["A790"] = FUTURE_POLICY

    # 5期見込（△△業ブロック R652付近を不動産用に転用）
    ws["A652"] = "今後５期の売上・営業利益見込み（不動産賃貸）"
    ws["C654"] = "第1期実績"
    ws["E654"] = "第2期見込"
    ws["G654"] = "第3期見込"
    ws["I654"] = "第4期見込"
    cols = ["C", "E", "G", "I"]
    ws["A656"] = "年間売上（千円）"
    ws["A657"] = "稼働率（%）"
    ws["A659"] = "年間売上（千円）"
    ws["A663"] = "年間営業利益（千円）"
    occupancy = [19.5, 77.4, 85, 90, 95]
    ws["C656"] = FORECAST[0]["sales_k"]
    ws["C657"] = occupancy[0]
    ws["C663"] = FORECAST[0]["op_k"]
    for i, col in enumerate(cols):
        f = FORECAST[i + 1]
        ws[f"{col}656"] = f["sales_k"]
        ws[f"{col}657"] = occupancy[i + 1]
        ws[f"{col}659"] = f["sales_k"]
        ws[f"{col}663"] = f["op_k"]


def main() -> None:
    for src, dst in [(TEMPL_H, OUT_H), (TEMPL_J, OUT_J)]:
        if not src.exists():
            raise FileNotFoundError(src)
        shutil.copy2(src, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb.active
        if dst == OUT_H:
            fill_hojin_rirekisho(ws)
        else:
            fill_jigyo_keikaku(ws)
        wb.save(dst)
        print(f"✅ 保存: {dst}")


if __name__ == "__main__":
    main()
