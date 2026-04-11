#!/usr/bin/env python3
"""
Vpass が出力する WEB 明細 CSV（複数カード区間が1ファイルに続く形式）を読み、
費目列を付けた税理士提出用 xlsx を出力する。仕分けルールは AMEX と同じ JSON を使う。

参照レイアウト: vpass_202509.xlsx（1行目メタ、2行目ヘッダ「日付・内容・金額・列1・費目」）

  python3 vpass_csv_classify.py preview ~/Downloads/202604.csv \\
    --rules ~/git-repos/215_kamiooya/C1_cursor/tax_docs_tools/amex_himoku_rules.json

  python3 vpass_csv_classify.py finalize ~/Downloads/202604.csv \\
    --rules .../amex_himoku_rules.json \\
    --output ~/Desktop/vpass_202604.xlsx

  # 費目が入っている行だけ（空欄行を落とす）
  python3 vpass_csv_classify.py export-tagged ~/Desktop/vpass_202604.xlsx \\
    --output ~/Desktop/vpass_202604_tagged.xlsx

複数 CSV を続けて渡すと、取引行を順に結合して1シートにまとめる（1行目メタは先頭ファイルのもの）。
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import tempfile
import os
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import numbers
except ImportError as e:
    raise SystemExit("openpyxl が必要です: pip install openpyxl") from e

from amex_activity_classify import classify_description, load_rules, normalize_text

DATE_HEAD = re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")
HEADERS = ("日付", "内容", "金額", "列1", "費目")
HIMOKU_COL = 5


def _is_amount_cell(s: str) -> bool:
    """WEB 明細 CSV の金額列（数値）。「１」等のフラグ列は除外。"""
    t = str(s).strip().replace(",", "")
    if not t or not re.match(r"^-?[\d.]+$", t):
        return False
    if t in ("1", "１"):
        return False
    try:
        float(t)
    except ValueError:
        return False
    return True


def extract_vpass_data_row(cells: list[str]) -> dict[str, str] | None:
    """
    1行分を date / desc / amount に正規化。
    Vpass の CSV は店名にカンマが含まれると列がずれるため、左から見て最初に現れる金額列を採用する。
    """
    if not cells:
        return None
    s0 = str(cells[0]).strip()
    if not DATE_HEAD.match(s0):
        return None
    rest = [str(c) if c is not None else "" for c in cells[1:]]
    # 通常: [内容, 金額, １, １, 金額, ...]
    if len(rest) >= 2 and _is_amount_cell(rest[1]):
        return {"date_str": s0, "desc": rest[0].strip(), "amount_raw": rest[1].strip()}
    amt_idx: int | None = None
    for i, cell in enumerate(rest):
        if _is_amount_cell(cell):
            amt_idx = i
            break
    if amt_idx is None or amt_idx < 1:
        return None
    desc = ", ".join(x.strip() for x in rest[:amt_idx]).strip()
    return {"date_str": s0, "desc": desc, "amount_raw": rest[amt_idx].strip()}


def decode_csv_bytes(raw: bytes) -> str:
    for enc in ("utf-8-sig", "cp932", "utf-8"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("cp932", errors="replace")


def parse_vpass_csv(path: Path) -> tuple[tuple[str, str, str] | None, list[dict[str, Any]]]:
    """
    戻り: (最初に見つかったメタ3列 or None, 取引行のリスト)
    各行 dict: date_str, desc, amount_raw, source_file
    """
    raw = path.read_bytes()
    text = decode_csv_bytes(raw)
    reader = csv.reader(text.splitlines())
    rows = list(reader)

    meta: tuple[str, str, str] | None = None
    out: list[dict[str, Any]] = []

    for row in rows:
        if not row:
            continue
        # 右側の空要素を整理しつつ最低3列
        row = [c if c is not None else "" for c in row]

        if len(row) >= 3 and "様" in str(row[0]) and "**" in str(row[1]):
            if meta is None:
                meta = (str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip())
            continue

        extracted = extract_vpass_data_row(row)
        if extracted is None:
            continue

        out.append(
            {
                **extracted,
                "source_file": path.name,
            }
        )

    return meta, out


def parse_date_for_excel(date_str: str) -> datetime | str:
    parts = date_str.split("/")
    if len(parts) != 3:
        return date_str
    y, mo, d = (int(parts[0]), int(parts[1]), int(parts[2]))
    try:
        return datetime(y, mo, d)
    except ValueError:
        return date_str


def amount_to_number(v: Any) -> int | float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if not s or not re.match(r"^-?[\d.]+$", s):
        return None
    if "." in s:
        return float(s)
    return int(s)


def run_preview(
    paths: list[Path],
    rules_path: Path,
    csv_out: Path | None,
    max_desc_width: int,
) -> int:
    rules_data = load_rules(rules_path)
    excludes = list(rules_data.get("exclude_row_substrings", []))
    himoku_rules = list(rules_data.get("himoku_rules", []))

    all_rows: list[dict[str, Any]] = []
    meta0: tuple[str, str, str] | None = None
    for p in paths:
        m, rows = parse_vpass_csv(p)
        if meta0 is None:
            meta0 = m
        all_rows.extend(rows)

    drops: list[tuple[str, str]] = []
    mapped: list[tuple[str, str, str]] = []
    blanks: list[tuple[str, str]] = []

    for row in all_rows:
        desc = row["desc"]
        cat = classify_description(str(desc), himoku_rules, excludes)
        day = row["date_str"]
        if cat is None:
            drops.append((day, normalize_text(desc)[:120]))
        elif cat == "":
            blanks.append((day, normalize_text(desc)[:200]))
        else:
            mapped.append((day, normalize_text(desc)[:max_desc_width], cat))

    print("=== Vpass 費目プレビュー（ルールは AMEX と共通）===")
    for p in paths:
        print(f"入力: {p}")
    print(f"ルール: {rules_path}")
    print(f"データ行: {len(all_rows)} / 除外(drop): {len(drops)} / 費目付与: {len(mapped)} / 未マッチ: {len(blanks)}")
    print()

    if drops:
        print("--- 除外される行（exclude_row_substrings）---")
        for day, d in drops:
            print(f"  {day}  {d}")
        print()

    if mapped:
        print("--- 仕分け結果（マッチした行）---")
        for day, d, h in mapped:
            print(f"  {day}  [{h}]  {d}")
        print()

    if blanks:
        print("--- 未マッチ（費目が空。ルールに contains を追加）---")
        for day, d in blanks:
            print(f"  {day}  {d}")
        print()

    if csv_out:
        csv_out.parent.mkdir(parents=True, exist_ok=True)
        with open(csv_out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["日付", "内容", "金額", "費目", "status", "source"])
            for row in all_rows:
                desc = str(row["desc"])
                amt = row["amount_raw"]
                day = row["date_str"]
                cat = classify_description(desc, himoku_rules, excludes)
                if cat is None:
                    st = "exclude"
                    cat_out = ""
                elif cat == "":
                    st = "unmapped"
                    cat_out = ""
                else:
                    st = "ok"
                    cat_out = cat
                w.writerow([day, desc, amt, cat_out, st, row["source_file"]])
        print(f"CSV を出力しました: {csv_out}")

    return 0


def run_finalize(
    paths: list[Path],
    rules_path: Path,
    output_path: Path,
    sheet_name: str,
    allow_unmapped: bool,
    drop_unmapped: bool,
) -> int:
    rules_data = load_rules(rules_path)
    excludes = list(rules_data.get("exclude_row_substrings", []))
    himoku_rules = list(rules_data.get("himoku_rules", []))

    all_rows: list[dict[str, Any]] = []
    meta0: tuple[str, str, str] | None = None
    for p in paths:
        m, rows = parse_vpass_csv(p)
        if meta0 is None:
            meta0 = m
        all_rows.extend(rows)

    unmapped: list[tuple[str, str]] = []
    for row in all_rows:
        cat = classify_description(str(row["desc"]), himoku_rules, excludes)
        if cat == "":
            unmapped.append((row["date_str"], normalize_text(row["desc"])[:100]))

    if unmapped and not allow_unmapped and not drop_unmapped:
        print("未マッチの行があります。プレビューでルールを足すか、次を指定してください:", file=sys.stderr)
        print("  --allow-unmapped … 費目空欄のまま出力", file=sys.stderr)
        print("  --drop-unmapped … 未マッチ行は出力しない", file=sys.stderr)
        for day, d in unmapped[:20]:
            print(f"  {day}: {d}", file=sys.stderr)
        if len(unmapped) > 20:
            print(f"  … 他 {len(unmapped) - 20} 件", file=sys.stderr)
        return 2

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)

    r = 1
    if meta0:
        for c, val in enumerate(meta0, start=1):
            ws.cell(r, c).value = val
        r += 1
    else:
        ws.cell(r, 1).value = "（メタ行なし）"
        r += 1

    for c, h in enumerate(HEADERS, start=1):
        ws.cell(r, c).value = h
    header_row = r
    r += 1

    for row in all_rows:
        desc = str(row["desc"])
        cat = classify_description(desc, himoku_rules, excludes)
        if cat is None:
            continue
        if cat == "" and drop_unmapped:
            continue
        if cat == "" and not allow_unmapped:
            continue

        dt = parse_date_for_excel(row["date_str"])
        ws.cell(r, 1).value = dt
        if isinstance(dt, datetime):
            ws.cell(r, 1).number_format = numbers.FORMAT_DATE_XLSX14
        ws.cell(r, 2).value = desc
        num = amount_to_number(row["amount_raw"])
        ws.cell(r, 3).value = num
        ws.cell(r, 4).value = None
        ws.cell(r, HIMOKU_COL).value = cat if cat else ""
        r += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"出力しました: {output_path} （シート: {sheet_name}）")
    return 0


def find_header_row_vpass(ws: Any) -> int:
    for r in range(1, min(ws.max_row, 30) + 1):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if v1 is not None and str(v1).strip() == "日付":
            if v2 is not None and str(v2).strip() == "内容":
                return r
    raise ValueError("「日付」「内容」のヘッダ行が見つかりません")


def run_export_tagged(input_path: Path, output_path: Path) -> int:
    wb_in = openpyxl.load_workbook(input_path, data_only=False)
    if len(wb_in.sheetnames) != 1:
        print("注意: 先頭シートのみ処理します", file=sys.stderr)
    ws = wb_in[wb_in.sheetnames[0]]
    header_row = find_header_row_vpass(ws)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    ws_new = wb_out.create_sheet(ws.title)

    for rr in range(1, header_row + 1):
        for c in range(1, 6):
            ws_new.cell(rr, c).value = ws.cell(rr, c).value

    out_r = header_row + 1
    kept = 0
    for r in range(header_row + 1, ws.max_row + 1):
        hv = ws.cell(r, HIMOKU_COL).value
        if hv is None or str(hv).strip() == "":
            continue
        for c in range(1, 6):
            ws_new.cell(out_r, c).value = ws.cell(r, c).value
        out_r += 1
        kept += 1

    wb_in.close()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if input_path.resolve() == output_path.resolve():
        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        tmp_path = Path(tmp)
        try:
            wb_out.save(tmp_path)
            tmp_path.replace(output_path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink(missing_ok=True)
    else:
        wb_out.save(output_path)
    print(f"費目のある行のみ {kept} 行を出力しました: {output_path}")
    return 0


def main() -> None:
    p = argparse.ArgumentParser(description="Vpass WEB明細 CSV の費目付与（AMEX と同じルール JSON）")
    sub = p.add_subparsers(dest="cmd", required=True)

    pp = sub.add_parser("preview", help="仕分け結果を表示")
    pp.add_argument("input", type=Path, nargs="+", help="202604.csv など（複数可）")
    pp.add_argument("--rules", type=Path, required=True, help="amex_himoku_rules.json（共通）")
    pp.add_argument("--csv", type=Path, help="レビュー用 CSV の出力先")
    pp.add_argument("--max-desc-width", type=int, default=60)

    pf = sub.add_parser("finalize", help="仕分け済み xlsx を出力（vpass_202509 相当レイアウト）")
    pf.add_argument("input", type=Path, nargs="+", help="WEB明細 CSV（複数は結合）")
    pf.add_argument("--rules", type=Path, required=True, help="amex_himoku_rules.json（共通）")
    pf.add_argument("--output", type=Path, required=True)
    pf.add_argument(
        "--sheet",
        type=str,
        default="",
        help="シート名（省略時は出力ファイル名の stem）",
    )
    pf.add_argument("--allow-unmapped", action="store_true", help="費目が空の行も残す")
    pf.add_argument("--drop-unmapped", action="store_true", help="費目が空の行は出力しない")

    pe = sub.add_parser("export-tagged", help="費目が入っている行だけ残す")
    pe.add_argument("input", type=Path, help="vpass finalize 済み xlsx")
    pe.add_argument("--output", type=Path, required=True)

    args = p.parse_args()
    if args.cmd == "preview":
        sys.exit(run_preview(args.input, args.rules, args.csv, args.max_desc_width))
    if args.cmd == "finalize":
        stem = args.output.stem
        sheet = args.sheet or stem or "vpass"
        sys.exit(
            run_finalize(
                args.input,
                args.rules,
                args.output,
                sheet,
                args.allow_unmapped,
                args.drop_unmapped,
            )
        )
    if args.cmd == "export-tagged":
        sys.exit(run_export_tagged(args.input, args.output))
    sys.exit(1)


if __name__ == "__main__":
    main()
