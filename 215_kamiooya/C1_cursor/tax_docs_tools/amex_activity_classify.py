#!/usr/bin/env python3
"""
AMEX の activity.xlsx（ご利用履歴シート）に費目列を付与し、除外行を落とす。

- preview: 仕分け結果を一覧表示（チャットで確認する前提）。未マッチ行を列挙。
- finalize: 確認後に税理士提出用ブックを出力（参考: amex_○月.xlsx と同様に「ご利用履歴」に費目列）。

使用例:

  python3 amex_activity_classify.py preview ~/Downloads/activity.xlsx \\
    --rules ~/path/to/amex_himoku_rules.json

  python3 amex_activity_classify.py finalize ~/Downloads/activity.xlsx \\
    --rules ~/path/to/amex_himoku_rules.json \\
    --output ~/Desktop/amex_仕分け済.xlsx

  # 費目が入っている行だけ残す（ブランク行削除）
  python3 amex_activity_classify.py export-tagged ~/Downloads/AMEX_4月.xlsx \\
    --output ~/Downloads/AMEX_4月.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import tempfile
import unicodedata
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    raise SystemExit("openpyxl が必要です: pip install openpyxl") from e

SHEET_MAIN = "ご利用履歴"
HIMOKU_HEADER = "費目"


def normalize_text(s: Any) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    t = t.replace("\u3000", " ")
    return re.sub(r"\s+", " ", t)


def load_rules(path: Path) -> dict[str, Any]:
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    data.pop("_comment", None)
    return data


def find_header_row(ws: Worksheet) -> int:
    for r in range(1, min(ws.max_row, 50) + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() == "ご利用日":
            return r
    raise ValueError("「ご利用日」行が見つかりません（ご利用履歴シートを確認）")


def row_is_empty_data(cells: list[Any]) -> bool:
    return not any(c is not None and str(c).strip() != "" for c in cells[:6])


def parse_data_rows(ws: Worksheet, header_row: int) -> list[tuple[int, dict[str, Any]]]:
    """(excel_row_index, row_dict) データ行。row_dict keys: col letters or names."""
    rows: list[tuple[int, dict[str, Any]]] = []
    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        headers.append("" if h is None else str(h).strip())

    for r in range(header_row + 1, ws.max_row + 1):
        cells = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if row_is_empty_data(cells):
            continue
        rowd = dict(zip(headers, cells))
        rowd["_row"] = r
        rows.append((r, rowd))
    return rows


def classify_description(
    desc: str,
    rules: list[dict[str, str]],
    excludes: list[str],
) -> str | None:
    """None = 行除外。'' = 未マッチ（費目空欄）。"""
    nd = normalize_text(desc)
    for ex in excludes:
        if ex and ex in nd:
            return None
    for rule in rules:
        sub = rule.get("contains", "")
        if sub and sub in nd:
            return rule["himoku"]
    return ""


def run_preview(
    input_path: Path,
    rules_path: Path,
    csv_out: Path | None,
    max_desc_width: int,
) -> int:
    rules_data = load_rules(rules_path)
    excludes = list(rules_data.get("exclude_row_substrings", []))
    himoku_rules = list(rules_data.get("himoku_rules", []))

    wb = openpyxl.load_workbook(input_path, data_only=True)
    if SHEET_MAIN not in wb.sheetnames:
        print(f"エラー: シート「{SHEET_MAIN}」がありません", file=sys.stderr)
        return 1
    ws = wb[SHEET_MAIN]
    header_row = find_header_row(ws)
    parsed = parse_data_rows(ws, header_row)

    drops: list[tuple[int, str, str]] = []
    mapped: list[tuple[int, str, str, str]] = []
    blanks: list[tuple[int, str, str]] = []

    for ridx, rowd in parsed:
        desc = rowd.get("ご利用内容", "") or ""
        amt = rowd.get("金額", "") or ""
        day = rowd.get("ご利用日", "") or ""
        cat = classify_description(str(desc), himoku_rules, excludes)
        if cat is None:
            drops.append((ridx, str(day), normalize_text(desc)[:120]))
        elif cat == "":
            blanks.append((ridx, str(day), normalize_text(desc)[:200]))
        else:
            mapped.append((ridx, str(day), normalize_text(desc)[:max_desc_width], cat))

    print("=== AMEX 費目プレビュー ===")
    print(f"入力: {input_path}")
    print(f"ルール: {rules_path}")
    print(f"データ行: {len(parsed)} / 除外(drop): {len(drops)} / 費目付与: {len(mapped)} / 未マッチ: {len(blanks)}")
    print()

    if drops:
        print("--- 除外される行（exclude_row_substrings）---")
        for ridx, day, d in drops:
            print(f"  行{ridx}  {day}  {d}")
        print()

    if mapped:
        print("--- 仕分け結果（マッチした行）---")
        for ridx, day, d, h in mapped:
            print(f"  行{ridx}  {day}  [{h}]  {d}")
        print()

    if blanks:
        print("--- 未マッチ（費目が空。ルールに contains を追加）---")
        for ridx, day, d in blanks:
            print(f"  行{ridx}  {day}  {d}")
        print()

    if csv_out:
        csv_out.parent.mkdir(parents=True, exist_ok=True)
        with open(csv_out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["excel_row", "ご利用日", "ご利用内容", "金額", "費目", "status"])
            for ridx, rowd in parsed:
                desc = str(rowd.get("ご利用内容", "") or "")
                amt = str(rowd.get("金額", "") or "")
                day = str(rowd.get("ご利用日", "") or "")
                cat = classify_description(desc, himoku_rules, excludes)
                if cat is None:
                    st = "exclude"
                    cat_out = ""
                elif cat == "":
                    st = "unmapped"
                    cat_out = ""
                else:
                    st = "ok"
                    cat_out = cat
                w.writerow([ridx, day, desc, amt, cat_out, st])
        print(f"CSV を出力しました: {csv_out}")

    wb.close()
    return 0


def run_finalize(
    input_path: Path,
    rules_path: Path,
    output_path: Path,
    allow_unmapped: bool,
    drop_unmapped: bool,
) -> int:
    rules_data = load_rules(rules_path)
    excludes = list(rules_data.get("exclude_row_substrings", []))
    himoku_rules = list(rules_data.get("himoku_rules", []))

    wb_in = openpyxl.load_workbook(input_path, data_only=False)
    if SHEET_MAIN not in wb_in.sheetnames:
        print(f"エラー: シート「{SHEET_MAIN}」がありません", file=sys.stderr)
        return 1

    ws = wb_in[SHEET_MAIN]
    header_row = find_header_row(ws)
    parsed = parse_data_rows(ws, header_row)

    # 未マッチチェック
    unmapped: list[tuple[int, str]] = []
    for ridx, rowd in parsed:
        desc = str(rowd.get("ご利用内容", "") or "")
        cat = classify_description(desc, himoku_rules, excludes)
        if cat == "":
            unmapped.append((ridx, normalize_text(desc)[:100]))

    if unmapped and not allow_unmapped and not drop_unmapped:
        print("未マッチの行があります。プレビューでルールを足すか、次を指定してください:", file=sys.stderr)
        print("  --allow-unmapped … 費目空欄のまま出力", file=sys.stderr)
        print("  --drop-unmapped … 未マッチ行は出力しない", file=sys.stderr)
        for ridx, d in unmapped[:20]:
            print(f"  行{ridx}: {d}", file=sys.stderr)
        if len(unmapped) > 20:
            print(f"  … 他 {len(unmapped) - 20} 件", file=sys.stderr)
        wb_in.close()
        return 2

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    ws_new = wb_out.create_sheet(SHEET_MAIN)

    # メタ行 + ヘッダー（費目列を追加）
    max_col_in = ws.max_column
    for r in range(1, header_row + 1):
        for c in range(1, max_col_in + 1):
            ws_new.cell(r, c).value = ws.cell(r, c).value
    ws_new.cell(header_row, max_col_in + 1).value = HIMOKU_HEADER

    out_row = header_row + 1
    for _ridx, rowd in parsed:
        desc = str(rowd.get("ご利用内容", "") or "")
        cat = classify_description(desc, himoku_rules, excludes)
        if cat is None:
            continue
        if cat == "" and drop_unmapped:
            continue
        if cat == "" and not allow_unmapped:
            continue

        src_row = rowd["_row"]
        for c in range(1, max_col_in + 1):
            ws_new.cell(out_row, c).value = ws.cell(src_row, c).value
        ws_new.cell(out_row, max_col_in + 1).value = cat if cat else ""
        out_row += 1

    # AMEX 提出用は「ご利用履歴」のみ（activity に付く「カードのご利用履歴」は出力しない）

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    wb_in.close()
    print(f"出力しました: {output_path}")
    return 0


def find_himoku_column(ws: Worksheet, header_row: int) -> int:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip() == HIMOKU_HEADER:
            return c
    raise ValueError("ヘッダー行に「費目」列がありません")


def run_export_tagged_only(input_path: Path, output_path: Path) -> int:
    """ご利用履歴で費目が空でないデータ行だけ残す（ヘッダー・メタ行は維持）。"""
    wb_in = openpyxl.load_workbook(input_path, data_only=False)
    if SHEET_MAIN not in wb_in.sheetnames:
        print(f"エラー: シート「{SHEET_MAIN}」がありません", file=sys.stderr)
        return 1

    ws = wb_in[SHEET_MAIN]
    header_row = find_header_row(ws)
    himoku_col = find_himoku_column(ws, header_row)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    ws_new = wb_out.create_sheet(SHEET_MAIN)

    max_col = ws.max_column
    for r in range(1, header_row + 1):
        for c in range(1, max_col + 1):
            ws_new.cell(r, c).value = ws.cell(r, c).value

    out_r = header_row + 1
    kept = 0
    for r in range(header_row + 1, ws.max_row + 1):
        hv = ws.cell(r, himoku_col).value
        if hv is None or str(hv).strip() == "":
            continue
        if row_is_empty_data([ws.cell(r, c).value for c in range(1, max_col + 1)]):
            continue
        for c in range(1, max_col + 1):
            ws_new.cell(out_r, c).value = ws.cell(r, c).value
        out_r += 1
        kept += 1

    # 「カードのご利用履歴」シートはコピーしない（提出用はご利用履歴のみ）

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_in.close()
    if input_path.resolve() == output_path.resolve():
        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        tmp_path = Path(tmp)
        try:
            wb_out.save(tmp_path)
            tmp_path.replace(output_path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink(missing_ok=True)
    else:
        wb_out.save(output_path)
    print(f"費目のある行のみ {kept} 行を出力しました: {output_path}")
    return 0


def main() -> None:
    p = argparse.ArgumentParser(description="AMEX activity.xlsx の費目付与・除外・出力")
    sub = p.add_subparsers(dest="cmd", required=True)

    pp = sub.add_parser("preview", help="仕分け結果を表示（チャット確認用）")
    pp.add_argument("input", type=Path, help="activity.xlsx など")
    pp.add_argument("--rules", type=Path, required=True, help="amex_himoku_rules.json")
    pp.add_argument("--csv", type=Path, help="レビュー用 CSV の出力先")
    pp.add_argument("--max-desc-width", type=int, default=60)

    pf = sub.add_parser("finalize", help="確認後に仕分け済み xlsx を出力")
    pf.add_argument("input", type=Path, help="activity.xlsx など")
    pf.add_argument("--rules", type=Path, required=True, help="amex_himoku_rules.json")
    pf.add_argument("--output", type=Path, required=True)
    pf.add_argument(
        "--allow-unmapped",
        action="store_true",
        help="費目が空の行も残す",
    )
    pf.add_argument(
        "--drop-unmapped",
        action="store_true",
        help="費目が空の行は出力しない",
    )

    pe = sub.add_parser(
        "export-tagged",
        help="費目が入っている行だけ残す（種目ブランクのデータ行を削除）",
    )
    pe.add_argument("input", type=Path, help="AMEX_○月.xlsx など（費目列付き）")
    pe.add_argument("--output", type=Path, required=True)

    args = p.parse_args()
    if args.cmd == "preview":
        sys.exit(run_preview(args.input, args.rules, args.csv, args.max_desc_width))
    if args.cmd == "finalize":
        sys.exit(
            run_finalize(
                args.input,
                args.rules,
                args.output,
                args.allow_unmapped,
                args.drop_unmapped,
            )
        )
    if args.cmd == "export-tagged":
        sys.exit(run_export_tagged_only(args.input, args.output))
    sys.exit(1)


if __name__ == "__main__":
    main()
