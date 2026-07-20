#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Notta / 動画文字起こし → knowledge_sources / knowledge_chunks

対応入力:
  - Notta Excel (.xlsx) … タイムスタンプ・話者付き想定
  - SRT (.srt)
  - 本文のみ Excel（旧サンプル）… 時刻なしチャンク

出力:
  - ローカル SQLite（knowledge_local.py）へ upsert
  - --csv で管理者互換＋動画メタCSVも出力可
  - SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY があれば Supabase へも upsert

使い方:
  python3 notta_to_knowledge.py --input path/to/file.xlsx --title "講義名" --dry-run
  python3 notta_to_knowledge.py --input path/to/file.srt --meta meta/notta_lessons.yaml --video-id step3
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from knowledge_local import (  # noqa: E402
    connect,
    counts,
    make_chunk_key,
    upsert_chunks,
    upsert_source,
)

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


@dataclass
class Cue:
    start_sec: int
    end_sec: int | None
    speaker: str | None
    text: str


_TS_RE = re.compile(
    r"^(?:(\d{1,2}):)?(\d{1,2}):(\d{2})(?:[.,](\d{1,3}))?$"
)
_SRT_TIME_RE = re.compile(
    r"(\d{2}):(\d{2}):(\d{2})[,.](\d{3})\s*-->\s*(\d{2}):(\d{2}):(\d{2})[,.](\d{3})"
)
_NOISE_RE = re.compile(r"^\[(音楽|拍手|沈黙|無音|Music|Applause)\]$", re.I)


def parse_timestamp(raw: str) -> int | None:
    s = (raw or "").strip()
    if not s:
        return None
    # already seconds
    if re.fullmatch(r"\d+", s):
        return int(s)
    m = _TS_RE.match(s.replace(" ", ""))
    if not m:
        return None
    h = int(m.group(1) or 0)
    mi = int(m.group(2))
    sec = int(m.group(3))
    return h * 3600 + mi * 60 + sec


def parse_srt(path: Path) -> list[Cue]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    blocks = re.split(r"\n\s*\n", text.strip())
    cues: list[Cue] = []
    for block in blocks:
        lines = [ln.strip("\ufeff") for ln in block.splitlines() if ln.strip()]
        if not lines:
            continue
        # optional index line
        if lines[0].isdigit() and len(lines) >= 2:
            lines = lines[1:]
        if not lines:
            continue
        m = _SRT_TIME_RE.search(lines[0])
        if not m:
            continue
        start = (
            int(m.group(1)) * 3600
            + int(m.group(2)) * 60
            + int(m.group(3))
        )
        end = (
            int(m.group(5)) * 3600
            + int(m.group(6)) * 60
            + int(m.group(7))
        )
        body_lines = lines[1:]
        speaker = None
        body: list[str] = []
        for bl in body_lines:
            if ":" in bl and len(bl.split(":", 1)[0]) <= 20:
                sp, rest = bl.split(":", 1)
                if rest.strip() and not sp.strip().isdigit():
                    speaker = sp.strip()
                    body.append(rest.strip())
                    continue
            body.append(bl)
        text_joined = " ".join(body).strip()
        if not text_joined or _NOISE_RE.match(text_joined):
            continue
        cues.append(Cue(start, end, speaker, text_joined))
    return cues


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", "", (h or "").strip().lower())


def parse_xlsx(path: Path) -> tuple[list[Cue], list[str]]:
    """Return cues and warnings."""
    if load_workbook is None:
        raise RuntimeError("openpyxl が必要です: pip install openpyxl")
    warnings: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["empty workbook"]

    # Detect header
    header_idx = 0
    headers = [str(c or "") for c in rows[0]]
    norm = [_normalize_header(h) for h in headers]
    header_like = any(
        n in ("timestamp", "timestamps", "開始", "開始時間", "時間", "time", "start", "話者", "speaker", "テキスト", "text", "内容", "transcription")
        for n in norm
    )
    if not header_like:
        # body-only: treat each non-empty cell in col B (or first non-empty) as a cue
        warnings.append("no_header_assumed_body_only")
        cues: list[Cue] = []
        for row in rows:
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            # skip sheet title-like short headers
            text = cells[-1] if len(cells) > 1 else cells[0]
            if _NOISE_RE.match(text):
                continue
            cues.append(Cue(0, None, None, text))
        return cues, warnings

    def col(*names: str) -> int | None:
        for i, n in enumerate(norm):
            for name in names:
                if name in n:
                    return i
        return None

    i_start = col("timestamp", "timestamps", "開始時間", "開始", "start", "time", "時間")
    i_end = col("end", "終了")
    i_speaker = col("speaker", "話者", "話者名")
    i_text = col("text", "transcription", "テキスト", "内容", "本文", "transcript")
    if i_text is None:
        # last column fallback
        i_text = len(headers) - 1
        warnings.append("text_column_fallback_last")

    # Merge full text heuristic: single huge row
    data_rows = rows[1:]
    if len(data_rows) <= 2 and i_start is None:
        warnings.append("possible_merged_full_text")

    cues = []
    for row in data_rows:
        vals = list(row) + [None] * (len(headers) - len(row))
        text = str(vals[i_text] or "").strip() if i_text is not None else ""
        if not text:
            continue
        if _NOISE_RE.match(text):
            continue
        start = parse_timestamp(str(vals[i_start] or "")) if i_start is not None else None
        end = parse_timestamp(str(vals[i_end] or "")) if i_end is not None else None
        speaker = str(vals[i_speaker] or "").strip() if i_speaker is not None else None
        if start is None:
            start = 0
            if i_start is not None:
                warnings.append("timestamp_parse_failed_row")
        cues.append(Cue(start, end, speaker or None, text))
    return cues, warnings


def chunk_cues(
    cues: list[Cue],
    *,
    max_sec: int = 40,
    max_chars: int = 600,
    overlap_sec: int = 8,
) -> list[dict[str, Any]]:
    if not cues:
        return []
    # If all start_sec==0 (body-only), pack by chars only
    body_only = all(c.start_sec == 0 for c in cues) and len(cues) > 1

    chunks: list[dict[str, Any]] = []
    buf: list[Cue] = []
    buf_start = 0
    buf_chars = 0

    def flush(next_overlap_from: int | None = None) -> None:
        nonlocal buf, buf_chars, buf_start
        if not buf:
            return
        text = " ".join(c.text for c in buf).strip()
        speakers = [c.speaker for c in buf if c.speaker]
        speaker = speakers[0] if speakers else None
        start = buf[0].start_sec
        end = buf[-1].end_sec if buf[-1].end_sec is not None else buf[-1].start_sec
        chunks.append(
            {
                "start_sec": start,
                "end_sec": end,
                "speaker": speaker,
                "content": text,
            }
        )
        if next_overlap_from is not None and not body_only and overlap_sec > 0:
            keep = [c for c in buf if c.start_sec >= next_overlap_from]
            buf = keep
            buf_chars = sum(len(c.text) for c in buf)
            buf_start = buf[0].start_sec if buf else 0
        else:
            buf = []
            buf_chars = 0
            buf_start = 0

    for cue in cues:
        if not buf:
            buf = [cue]
            buf_start = cue.start_sec
            buf_chars = len(cue.text)
            continue
        span = cue.start_sec - buf_start
        would_chars = buf_chars + 1 + len(cue.text)
        if (not body_only and span >= max_sec) or would_chars >= max_chars:
            overlap_from = max(0, cue.start_sec - overlap_sec)
            flush(next_overlap_from=overlap_from if not body_only else None)
            if not buf:
                buf = [cue]
                buf_start = cue.start_sec
                buf_chars = len(cue.text)
            else:
                buf.append(cue)
                buf_chars += 1 + len(cue.text)
        else:
            buf.append(cue)
            buf_chars = would_chars
    flush()
    return chunks


def load_meta(path: Path | None, video_id: str | None) -> dict[str, Any]:
    if not path:
        return {}
    if yaml is None:
        raise RuntimeError("PyYAML が必要です: pip install pyyaml")
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    lessons = data.get("lessons") or data
    if isinstance(lessons, dict) and video_id and video_id in lessons:
        return lessons[video_id] or {}
    if isinstance(lessons, list) and video_id:
        for item in lessons:
            if str(item.get("video_id") or item.get("id")) == video_id:
                return item
    return data if isinstance(data, dict) and "title" in data else {}


def supabase_upsert(source: dict, chunks: list[dict]) -> tuple[int, int]:
    url = (os.environ.get("SUPABASE_URL") or "").strip().rstrip("/")
    key = (
        (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
        or (os.environ.get("SUPABASE_ANON_KEY") or "").strip()
    )
    if not url or not key:
        return 0, 0
    try:
        from supabase import create_client
    except ImportError:
        print("[WARN] supabase package missing; skip remote upsert", file=sys.stderr)
        return 0, 0
    try:
        client = create_client(url, key)
        client.table("knowledge_sources").upsert(source, on_conflict="source_key").execute()
        # resolve source id
        res = (
            client.table("knowledge_sources")
            .select("id")
            .eq("source_key", source["source_key"])
            .limit(1)
            .execute()
        )
        sid = res.data[0]["id"] if res.data else None
        if sid is None:
            return 1, 0
        payload = []
        for c in chunks:
            payload.append(
                {
                    "source_id": sid,
                    "chunk_key": c["chunk_key"],
                    "start_sec": c["start_sec"],
                    "end_sec": c.get("end_sec"),
                    "speaker": c.get("speaker"),
                    "content": c["content"],
                    "content_hash": c["content_hash"],
                    "search_text": f'{c.get("speaker") or ""} {c["content"]}'.strip(),
                }
            )
        for i in range(0, len(payload), 200):
            client.table("knowledge_chunks").upsert(
                payload[i : i + 200], on_conflict="chunk_key"
            ).execute()
        return 1, len(payload)
    except Exception as e:
        print(f"[WARN] Supabase upsert failed: {e}", file=sys.stderr)
        return 0, 0


def parse_input(path: Path) -> tuple[list[Cue], list[str]]:
    suf = path.suffix.lower()
    if suf == ".srt":
        return parse_srt(path), []
    if suf in (".xlsx", ".xlsm"):
        return parse_xlsx(path)
    raise RuntimeError(f"未対応形式: {suf}")


def main() -> int:
    ap = argparse.ArgumentParser(description="Notta/SRT → knowledge chunks")
    ap.add_argument("--input", "-i", required=True, help="xlsx or srt")
    ap.add_argument("--title", default="", help="動画タイトル")
    ap.add_argument("--video-id", default="", help="安定 video_id")
    ap.add_argument("--video-url", default="", help="動画URL")
    ap.add_argument("--instructor", default="", help="講師名")
    ap.add_argument("--offset-sec", type=int, default=0, help="動画内開始オフセット秒")
    ap.add_argument("--meta", default="", help="YAML メタファイル")
    ap.add_argument("--csv", default="", help="出力CSVパス")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-local", action="store_true")
    ap.add_argument("--skip-supabase", action="store_true")
    args = ap.parse_args()

    path = Path(args.input).expanduser().resolve()
    if not path.is_file():
        print(f"入力がありません: {path}", file=sys.stderr)
        return 2

    meta = load_meta(Path(args.meta) if args.meta else None, args.video_id or None)
    title = args.title or meta.get("title") or path.stem
    video_id = args.video_id or meta.get("video_id") or path.stem
    video_url = args.video_url or meta.get("video_url") or ""
    instructor = args.instructor or meta.get("instructor") or ""
    offset = int(args.offset_sec or meta.get("offset_sec") or 0)

    cues, warnings = parse_input(path)
    if offset:
        for c in cues:
            c.start_sec = max(0, c.start_sec + offset)
            if c.end_sec is not None:
                c.end_sec = max(0, c.end_sec + offset)

    raw_chunks = chunk_cues(cues)
    chunks = []
    for ch in raw_chunks:
        text = ch["content"]
        start = int(ch["start_sec"])
        h = hashlib.sha1(text.encode("utf-8")).hexdigest()[:16]
        chunks.append(
            {
                **ch,
                "video_id": video_id,
                "content_hash": h,
                "chunk_key": make_chunk_key(video_id, start, text),
            }
        )

    source_key = f"notta:{video_id}"
    print(
        f"input={path.name} cues={len(cues)} chunks={len(chunks)} "
        f"title={title!r} video_id={video_id} warnings={warnings}"
    )

    if args.csv:
        out = Path(args.csv).expanduser().resolve()
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(
                f,
                fieldnames=[
                    "chunk_key",
                    "video_id",
                    "video_title",
                    "start_sec",
                    "end_sec",
                    "speaker",
                    "content",
                    "video_url",
                    "ソース",
                ],
            )
            w.writeheader()
            for ch in chunks:
                w.writerow(
                    {
                        "chunk_key": ch["chunk_key"],
                        "video_id": video_id,
                        "video_title": title,
                        "start_sec": ch["start_sec"],
                        "end_sec": ch.get("end_sec") or "",
                        "speaker": ch.get("speaker") or "",
                        "content": ch["content"],
                        "video_url": video_url,
                        "ソース": "WeStudyセミナー動画",
                    }
                )
        print(f"csv: {out}")

    if args.dry_run:
        for ch in chunks[:3]:
            print(" sample:", ch["start_sec"], ch["content"][:60])
        return 0

    if not args.skip_local:
        conn = connect()
        sid = upsert_source(
            conn,
            source_key=source_key,
            title=title,
            video_id=video_id,
            video_url=video_url or None,
            instructor=instructor or None,
            origin_path=str(path),
            meta={
                "warnings": warnings,
                "offset_sec": offset,
                "origin_filename": path.name,
                "brand": "WeStudy",
            },
            content_channel="seminar_video",
        )
        n = upsert_chunks(conn, sid, chunks)
        print(f"local upsert: source_id={sid} chunks={n} counts={counts(conn)}")

    if not args.skip_supabase:
        src_row = {
            "source_key": source_key,
            "source_kind": "video",
            "content_channel": "seminar_video",
            "title": title,
            "video_id": video_id,
            "video_url": video_url or None,
            "instructor": instructor or None,
            "origin_path": str(path),
            "meta_json": {
                "warnings": warnings,
                "offset_sec": offset,
                "origin_filename": path.name,
                "brand": "WeStudy",
            },
            "ingest_status": "ready",
        }
        s_n, c_n = supabase_upsert(src_row, chunks)
        print(f"supabase upsert: sources={s_n} chunks={c_n}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
