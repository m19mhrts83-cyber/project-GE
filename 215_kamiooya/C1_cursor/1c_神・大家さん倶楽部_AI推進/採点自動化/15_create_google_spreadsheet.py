#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""WeStudy 採点自動化 — xlsx から Google スプレッドシートを新規作成する。

使い方:
  cd ~/git-repos/215_kamiooya/C1_cursor/1c_神・大家さん倶楽部_AI推進/採点自動化
  /Users/matsunomasaharu2/selenium_env/venv/bin/python 15_create_google_spreadsheet.py \\
    --source-xlsx ~/Downloads/WeStudy_採点自動化.xlsx

  # 設定 YAML に spreadsheet ID を書き込む
  ... --write-config

  # 作成のみ（dry-run 相当: シート構成だけ確認）
  ... --dry-run
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
WESTUDY_COMMON = SCRIPT_DIR.parent.parent / "westudy_common"
if str(WESTUDY_COMMON) not in sys.path:
    sys.path.insert(0, str(WESTUDY_COMMON))

from googleapiclient.discovery import build  # noqa: E402

from google_workspace_auth import load_credentials  # noqa: E402

CONFIG_PATH = WESTUDY_COMMON / "kamiooya_google_config.yaml"

SHEET_NAMES = [
    "設定",
    "元データ",
    "得点基準",
    "採点結果",
    "集計",
    "ログ",
    "補正学習データ",
    "バージョン履歴",
]

RESULT_HEADERS = [
    "コメントID",
    "投稿日時",
    "投稿者名",
    "親コメントID",
    "コメント内容",
    "対象判定",
    "推定分野",
    "サブ分類",
    "ルールID",
    "得点",
    "根拠",
    "根拠抜粋",
    "信頼度",
    "手動補正点",
    "最終点",
    "採点バージョン",
    "採点日時",
    "エラー",
]

LOG_HEADERS = ["時刻", "レベル", "関数名", "コメントID", "メッセージ"]

LEARN_HEADERS = [
    "追加日時",
    "コメントID",
    "投稿者名",
    "コメント内容",
    "Gemini得点",
    "手動補正点",
    "点差",
    "GeminiルールID",
    "最終ルールID",
    "補正理由メモ",
    "採点バージョン",
]

AGGREGATE_HEADERS = [
    "投稿者名",
    "投稿数",
    "合計点",
    "平均点",
    "備考",
]

SETTING_KEYS = [
    "GEMINI_API_KEY",
    "GEMINI_MODEL",
    "TARGET_MONTH",
    "DRIVE_CSV_FOLDER_ID",
    "DRIVE_CSV_FILENAME",
    "SOURCE_SHEET_NAME",
    "RULES_SHEET_NAME",
    "RESULT_SHEET_NAME",
    "LOG_SHEET_NAME",
    "MAX_ROWS_PER_RUN",
    "INCLUDE_REPLIES",
]


def _load_xlsx(path: Path) -> dict[str, list[list]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[str, list[list]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            rows.append(list(row))
        out[name] = rows
    return out


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (dt.datetime, dt.date)):
        if isinstance(v, dt.datetime):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _normalize_setting_rows(raw_rows: list[list]) -> list[list]:
    """xlsx の設定シートを key-value 形式に正規化。"""
    by_key: dict[str, str] = {}
    for row in raw_rows:
        if len(row) < 2:
            continue
        k = _cell_str(row[0])
        if not k:
            continue
        by_key[k] = _cell_str(row[1])

    # 正規化ルール
    if "TARGET_MONTH" in by_key:
        m = re.match(r"(\d{4})-(\d{1,2})", by_key["TARGET_MONTH"])
        if m:
            by_key["TARGET_MONTH"] = f"{m.group(1)}-{int(m.group(2)):02d}"
        elif re.match(r"\d{4}-\d{2}-\d{2}", by_key["TARGET_MONTH"]):
            d = dt.datetime.strptime(by_key["TARGET_MONTH"][:10], "%Y-%m-%d")
            by_key["TARGET_MONTH"] = d.strftime("%Y-%m")

    defaults = {
        "GEMINI_MODEL": "gemini-2.5-flash",
        "TARGET_MONTH": "",
        "DRIVE_CSV_FOLDER_ID": "1QI-r0upkP335FZNQ8q99pe7FpPfTt1h_",
        "DRIVE_CSV_FILENAME": "WeStudy_for_scoring.csv",
        "SOURCE_SHEET_NAME": "元データ",
        "RULES_SHEET_NAME": "得点基準",
        "RESULT_SHEET_NAME": "採点結果",
        "LOG_SHEET_NAME": "ログ",
        "MAX_ROWS_PER_RUN": "5",
        "INCLUDE_REPLIES": "FALSE",
    }
    for k, v in defaults.items():
        by_key.setdefault(k, v)

    # 試行用: MAX_ROWS は 5 固定（xlsx が 50 でも上書きしない場合は xlsx 優先）
    if not by_key.get("MAX_ROWS_PER_RUN"):
        by_key["MAX_ROWS_PER_RUN"] = "5"

    return [[k, by_key.get(k, "")] for k in SETTING_KEYS if k in by_key or k in defaults]


def _sheet_rows_for_upload(name: str, xlsx_data: dict[str, list[list]]) -> list[list]:
    if name == "設定":
        raw = xlsx_data.get("設定", [])
        return _normalize_setting_rows(raw)
    if name == "元データ":
        return [[_cell_str(c) for c in row] for row in xlsx_data.get("元データ", [])]
    if name == "得点基準":
        rows = xlsx_data.get("得点基準", [])
        out = []
        for row in rows:
            cells = []
            for c in row:
                if isinstance(c, bool):
                    cells.append("TRUE" if c else "FALSE")
                else:
                    cells.append(_cell_str(c))
            out.append(cells)
        return out
    if name == "採点結果":
        return [RESULT_HEADERS]
    if name == "ログ":
        return [LOG_HEADERS]
    if name == "補正学習データ":
        return [LEARN_HEADERS]
    if name == "集計":
        return [AGGREGATE_HEADERS]
    return []


def _create_spreadsheet(sheets, title: str) -> tuple[str, dict[str, int]]:
    """新規 SS 作成。sheetId マップを返す。"""
    body = {
        "properties": {"title": title},
        "sheets": [{"properties": {"title": n}} for n in SHEET_NAMES],
    }
    resp = sheets.spreadsheets().create(body=body, fields="spreadsheetId,sheets.properties").execute()
    ss_id = resp["spreadsheetId"]
    name_to_id = {
        sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in resp.get("sheets", [])
    }
    return ss_id, name_to_id


def _upload_values(sheets, ss_id: str, sheet_name: str, rows: list[list]) -> None:
    if not rows:
        return
    sheets.spreadsheets().values().update(
        spreadsheetId=ss_id,
        range=f"'{sheet_name}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows},
    ).execute()


def _write_config_spreadsheet_id(ss_id: str) -> None:
    import yaml

    cfg: dict = {}
    if CONFIG_PATH.is_file():
        cfg = yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}
    cfg["saiten_spreadsheet_id"] = ss_id
    CONFIG_PATH.write_text(
        yaml.safe_dump(cfg, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    print(f"📎 kamiooya_google_config.yaml に saiten_spreadsheet_id を保存しました", file=sys.stderr)


def main() -> int:
    parser = argparse.ArgumentParser(description="採点自動化 Google スプレッドシート作成")
    parser.add_argument(
        "--source-xlsx",
        type=Path,
        default=Path.home() / "Downloads/WeStudy_採点自動化.xlsx",
        help="移行元 xlsx",
    )
    parser.add_argument("--title", default="WeStudy_採点自動化", help="スプレッドシート名")
    parser.add_argument("--write-config", action="store_true", help="YAML に spreadsheet ID を保存")
    parser.add_argument("--dry-run", action="store_true", help="xlsx 読取のみ（API 呼び出しなし）")
    args = parser.parse_args()

    if not args.source_xlsx.is_file():
        print(f"❌ xlsx が見つかりません: {args.source_xlsx}", file=sys.stderr)
        return 1

    xlsx_data = _load_xlsx(args.source_xlsx)
    print(f"✅ xlsx 読込: {args.source_xlsx.name}", file=sys.stderr)
    for sn in ("設定", "元データ", "得点基準"):
        n = len(xlsx_data.get(sn, []))
        print(f"   {sn}: {n} 行", file=sys.stderr)

    if args.dry_run:
        print("dry-run のため Google API は呼び出しません", file=sys.stderr)
        return 0

    creds = load_credentials()
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)

    ss_id, _ = _create_spreadsheet(sheets, args.title)
    url = f"https://docs.google.com/spreadsheets/d/{ss_id}/edit"

    for sheet_name in SHEET_NAMES:
        rows = _sheet_rows_for_upload(sheet_name, xlsx_data)
        _upload_values(sheets, ss_id, sheet_name, rows)
        print(f"   投入: {sheet_name} ({len(rows)} 行)", file=sys.stderr)

    print("")
    print("✅ Google スプレッドシートを作成しました")
    print(f"   タイトル: {args.title}")
    print(f"   ID: {ss_id}")
    print(f"   URL: {url}")
    print("")
    print("次の手順:")
    print("  1. 上記 URL を開く")
    print("  2. 拡張機能 > Apps Script → 04_gas_Code.gs を Code.gs に貼付")
    print("  3. 保存 → シート再読込 → メニュー「採点自動化」を確認")
    print("  4. initializeSheets は不要（ヘッダー済み）。初回 runScoring で権限許可")

    if args.write_config:
        _write_config_spreadsheet_id(ss_id)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
