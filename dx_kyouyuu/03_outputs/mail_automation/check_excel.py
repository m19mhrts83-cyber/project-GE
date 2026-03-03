#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

# Excelファイルを読み込む
wb = openpyxl.load_workbook("テスト送信リスト.xlsx", data_only=True)
ws = wb.active

print("=" * 60)
print("📊 送信先リストの内容")
print("=" * 60)
print()

# すべての行を表示
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row_idx == 1:
        # ヘッダー行
        print(f"【ヘッダー】")
        print(" | ".join([str(cell) if cell else "" for cell in row]))
        print("-" * 60)
    else:
        # データ行
        if any(cell for cell in row):  # 空行でない場合
            print(f"{row_idx-1}. ", end="")
            for cell in row:
                if cell:
                    print(f"{cell} | ", end="")
            print()

print()
print("=" * 60)

# メールアドレスのカウント
email_col_idx = None
for idx, cell in enumerate(ws[1]):
    if cell.value and 'メール' in str(cell.value).lower():
        email_col_idx = idx
        break

if email_col_idx is not None:
    email_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[email_col_idx] and '@' in str(row[email_col_idx]):
            email_count += 1
    
    print(f"✅ メールアドレス総数: {email_count}件")
else:
    print("⚠️ メールアドレス列が見つかりませんでした")

print("=" * 60)
