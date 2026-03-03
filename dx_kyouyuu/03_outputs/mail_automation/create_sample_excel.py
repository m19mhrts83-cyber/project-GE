#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
メールアドレス一覧のサンプルExcelファイルを作成するスクリプト
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def create_sample_excel():
    """サンプルのExcelファイルを作成する"""
    
    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "送信先一覧"
    
    # ヘッダー行のスタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ヘッダー行を作成
    headers = ["会社名", "担当者", "メールアドレス", "備考"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # サンプルデータ
    sample_data = [
        ["A不動産", "田中様", "tanaka@example.com", ""],
        ["B管理", "鈴木様", "suzuki@example.com", ""],
        ["C仲介", "佐藤様", "sato@example.com", ""],
        ["D賃貸", "高橋様", "takahashi@example.com", ""],
        ["E物件", "渡辺様", "watanabe@example.com", ""],
    ]
    
    # データ行を追加
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="center")
    
    # 列幅を自動調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    
    # ファイルを保存
    output_path = Path(__file__).parent / "送信先メールアドレス一覧_サンプル.xlsx"
    wb.save(output_path)
    print(f"サンプルExcelファイルを作成しました: {output_path}")
    print("\nこのファイルを編集して、実際の送信先メールアドレスを入力してください。")
    print("注意: example.comのアドレスは実在しないため、実際の送信には使用できません。")

if __name__ == '__main__':
    create_sample_excel()
