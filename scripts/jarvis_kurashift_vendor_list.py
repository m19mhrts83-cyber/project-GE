#!/usr/bin/env python3
"""地場不動産 — 業者開拓リストの取込・状態更新・Grok 日次探索用出力。

正本: config/kurashift_re_vendor_list.yaml（.gitignore。example/template を参照）

  cd ~/git-repos
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --summary
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --import-csv path/to/list.csv
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --next 3
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --batch-week --grok-kickoff
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --apply-marks grok_summary.txt
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --mark ID --status contacted --note "Web送信"
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --grok-discovery-prompt
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_vendor_list.py --merge-append block.yaml
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import yaml

REPO = Path(__file__).resolve().parents[1]
LIST_PATH = REPO / "config" / "kurashift_re_vendor_list.yaml"
EXAMPLE_PATH = REPO / "config" / "kurashift_re_vendor_list.example.yaml"
TEMPLATE_CSV = REPO / "config" / "kurashift_re_vendor_list.template.csv"
GROK_APPEND_DOC = REPO / "config" / "grok_vendor_discovery_append.md"

STATUSES = frozenset(
    {"pending", "contacted", "replied", "skip", "discovered", "invalid"}
)

REGION_ORDER = {"chubu": 0, "shiga": 1, "list": 9}

PHASE_DAILY_LIMIT = {1: 3, 2: 5, 3: 10}

JARVIS_PRIVATE = REPO / ".env.jarvis_private"

MARK_CMD_RE = re.compile(
    r"--mark\s+(?P<id>[^\s]+)\s+--status\s+(?P<status>[^\s]+)"
    r'(?:\s+--note\s+"(?P<note_dq>(?:[^"\\]|\\.)*)"|'
    r"\s+--note\s+(?P<note_sq>[^\s]+))?",
    re.I,
)

DISCOVERED_URL_RE = re.compile(
    r"discovered_url:\s*(https?://[^\s|\"'<>]+)",
    re.I,
)

CHAIN_PATTERNS: tuple[tuple[str, str], ...] = (
    ("daikyo-anabuki", r"大京穴吹|daikyo[\-\.]?anabuki"),
    ("pitathouse", r"ピタット|PITATHOUSE|pitathouse"),
    ("century21", r"センチュリー\s*21|CENTURY\s*21"),
    ("tokyu-livable", r"東急リバブル|livable"),
    ("sumitomo-fudosan", r"住友不動産"),
    ("mitsui-rehouse", r"三井のリハウス|リハウス"),
    ("suumo", r"スーモ|SUUMO"),
    ("athome", r"アットホーム|athome"),
    ("o-uccino", r"オウチーノ|ouchi"),
)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def slug_id(name: str, area: str = "") -> str:
    base = f"{name}-{area}".strip().lower()
    base = re.sub(r"\s+", "-", base)
    base = re.sub(r"[^a-z0-9\-ぁ-んァ-ヶ一-龥]", "", base)
    if len(base) >= 8:
        return base[:48]
    h = hashlib.sha256(base.encode()).hexdigest()[:8]
    return f"v-{h}"


def load_list(path: Path | None = None) -> dict[str, Any]:
    p = path or LIST_PATH
    if not p.is_file():
        if EXAMPLE_PATH.is_file():
            data = yaml.safe_load(EXAMPLE_PATH.read_text(encoding="utf-8")) or {}
            data.setdefault("settings", {})
            data.setdefault("vendors", [])
            return data
        return {"settings": {}, "vendors": []}
    return yaml.safe_load(p.read_text(encoding="utf-8")) or {"settings": {}, "vendors": []}


def save_list(data: dict[str, Any], path: Path | None = None) -> None:
    p = path or LIST_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def vendor_index(data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for v in data.get("vendors") or []:
        if isinstance(v, dict) and v.get("id"):
            out[str(v["id"])] = v
    return out


def row_to_vendor(row: dict[str, str], *, source: str = "import") -> dict[str, Any]:
    name = (row.get("name") or "").strip()
    area = (row.get("area") or "").strip()
    vid = (row.get("id") or "").strip() or slug_id(name, area)
    status = (row.get("status") or "pending").strip() or "pending"
    if status not in STATUSES:
        status = "pending"
    return {
        "id": vid,
        "name": name,
        "area": area,
        "prefecture": (row.get("prefecture") or "").strip(),
        "city": (row.get("city") or "").strip(),
        "url": (row.get("url") or "").strip(),
        "contact_url": (row.get("contact_url") or row.get("contact") or "").strip(),
        "channel": (row.get("channel") or "web_form").strip() or "web_form",
        "contact_email": (row.get("contact_email") or row.get("email") or "").strip(),
        "phone": (row.get("phone") or "").strip(),
        "status": status,
        "source": (row.get("source") or source).strip() or source,
        "notes": (row.get("notes") or "").strip(),
        "discovered_at": (row.get("discovered_at") or "").strip(),
        "contacted_at": (row.get("contacted_at") or "").strip(),
        "replied_at": (row.get("replied_at") or "").strip(),
        "last_result": (row.get("last_result") or "").strip(),
    }


def _prefecture_label(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    if s.endswith("県"):
        return s
    return f"{s}県"


def _excel_cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _cell_idx(row: tuple[Any, ...], idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return _excel_cell_str(row[idx])


def _region_from_path(path: Path) -> tuple[str, str]:
    s = path.as_posix()
    if "滋賀" in s:
        return "shiga", "kamiooya_滋賀"
    if "中部" in s or "東海" in s:
        return "chubu", "kamiooya_中部東海"
    return "list", "kamiooya_import"


def _detect_xlsx_format(headers: tuple[Any, ...]) -> str:
    form_cols = [
        i for i, h in enumerate(headers) if h and "問い合わせフォーム" in str(h)
    ]
    if len(form_cols) >= 2:
        return "shiga_ops"
    return "standard"


def import_kamiooya_xlsx(path: Path, *, dry_run: bool, merge: bool = False) -> dict[str, Any]:
    """神大家「地場不動産業者一覧」Excel を取込（中部東海 / 滋賀 等）。"""
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl required for --import-xlsx") from exc

    list_region, source_tag = _region_from_path(path)
    id_width = 4 if list_region == "shiga" else 3

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    headers: tuple[Any, ...] = ()
    for i, row in enumerate(rows):
        if row and row[0] == "No.":
            header_idx = i
            headers = row
            break
    if header_idx is None:
        return {"ok": False, "error": "header row (No.) not found"}

    xlsx_format = _detect_xlsx_format(headers)

    data = load_list()
    settings = data.setdefault("settings", {})
    if not merge:
        data["vendors"] = []
    settings.setdefault("daily_outreach_limit", 3)
    settings.setdefault("daily_discovery_limit", 5)
    sources = settings.get("source_xlsx") or []
    if isinstance(sources, str):
        sources = [sources]
    if str(path) not in sources:
        sources.append(str(path))
    settings["source_xlsx"] = sources
    settings["imported_at"] = now_iso()[:10]
    settings["outreach_from"] = "matsuno.estate@gmail.com"
    settings["outreach_identity"] = "personal"
    settings["outreach_note"] = (
        "送信は松野個人・松野エステイト（matsuno.estate）。"
        "滋賀リストの運営問合済みは参考のみ。個人として改めて問合せする。"
    )
    areas = set(settings.get("target_areas") or [])
    areas.update(["愛知県", "岐阜県", "静岡県", "三重県", "滋賀県"])
    settings["target_areas"] = sorted(areas)

    by_id = vendor_index(data)
    added = updated = skipped = 0

    for row in rows[header_idx + 1 :]:
        if not row or not row[0]:
            continue
        no_raw = _cell_idx(row, 0)
        if no_raw == "記入例" or not no_raw.isdigit():
            continue

        name = _cell_idx(row, 4)
        if not name:
            skipped += 1
            continue

        pref = _prefecture_label(_cell_idx(row, 1))
        city = _cell_idx(row, 2)
        station = _cell_idx(row, 3)
        prop_area = _cell_idx(row, 5)
        url = _cell_idx(row, 6)
        if url and not url.startswith("http"):
            url = ""

        remark = _cell_idx(row, 9 if xlsx_format == "shiga_ops" else 7)
        mail = _cell_idx(row, 17 if xlsx_format == "shiga_ops" else 15)
        tel = _cell_idx(row, 18 if xlsx_format == "shiga_ops" else 16)

        form_has = ""
        ops_contacted_at = ""
        status = "pending"
        form_note = ""

        if xlsx_format == "shiga_ops":
            raw_form7 = row[7] if len(row) > 7 else None
            form_has = _cell_idx(row, 16)
            if isinstance(raw_form7, datetime):
                # 運営作業の問合日。個人（松野・estate）未送信のため status は pending のまま
                ops_contacted_at = raw_form7.strftime("%Y-%m-%d")
                form_note = f"運営問合済(参考):{ops_contacted_at}"
            elif raw_form7 is not None:
                s7 = _excel_cell_str(raw_form7)
                if s7 and s7 not in ("なし", ""):
                    form_note = f"問合せ欄7:{s7}"
        else:
            form_has = _cell_idx(row, 14)

        if form_has == "有り":
            channel = "web_form"
            contact_url = url
        elif mail:
            channel = "email"
            contact_url = ""
        elif tel:
            channel = "phone"
            contact_url = ""
        else:
            channel = "web_form"
            contact_url = ""

        area_parts = [p for p in (pref, city) if p]
        area = "".join(area_parts) if area_parts else prop_area or pref

        note_bits: list[str] = []
        if station:
            note_bits.append(f"最寄: {station}")
        if prop_area:
            note_bits.append(f"物件エリア: {prop_area}")
        if remark:
            note_bits.append(remark)
        if form_note:
            note_bits.append(form_note)
        for label, idx in (
            ("Line/メルマガ", 11),
            ("会員登録", 12),
            ("売却査定", 10),
            ("問合せフォーム", 16 if xlsx_format == "shiga_ops" else 14),
            ("HP", 15),
        ):
            v = _cell_idx(row, idx)
            if v and v not in ("なし", " "):
                note_bits.append(f"{label}:{v}")

        reg_memo = _cell_idx(row, 20)
        if reg_memo:
            note_bits.append(f"作業メモ:{reg_memo}")

        reply_greet = _cell_idx(row, 21)
        reply_prop = _cell_idx(row, 22)
        if reply_prop:
            status = "replied"
        elif reply_greet and status == "pending":
            status = "contacted"

        vid = f"{list_region}-{int(no_raw):0{id_width}d}"

        vendor = {
            "id": vid,
            "name": name,
            "area": area,
            "prefecture": pref,
            "city": city,
            "url": url,
            "contact_url": contact_url if form_has == "有り" and url else "",
            "channel": channel,
            "contact_email": mail,
            "phone": tel,
            "status": status,
            "source": source_tag,
            "notes": " / ".join(note_bits),
            "discovered_at": now_iso()[:10],
            "list_no": int(no_raw),
            "list_region": list_region,
        }
        if ops_contacted_at:
            vendor["ops_contacted_at"] = ops_contacted_at

        existing = by_id.get(vid)
        if existing:
            for k, val in vendor.items():
                if k == "id":
                    continue
                if val or k in ("status", "contacted_at", "ops_contacted_at"):
                    existing[k] = val
            if ops_contacted_at:
                existing["status"] = "pending"
                existing.pop("contacted_at", None)
            updated += 1
        else:
            data.setdefault("vendors", []).append(vendor)
            by_id[vid] = vendor
            added += 1

    if not dry_run:
        save_list(data)

    return {
        "ok": True,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "total": len(data.get("vendors") or []),
        "path": str(LIST_PATH),
        "source": str(path),
        "list_region": list_region,
        "merge": merge,
        "dry_run": dry_run,
    }


def import_csv(path: Path, *, dry_run: bool) -> dict[str, Any]:
    data = load_list()
    by_id = vendor_index(data)
    added = updated = skipped = 0
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("name") or "").strip()
            if not name or name.startswith("（例）"):
                skipped += 1
                continue
            v = row_to_vendor(row, source="import_csv")
            if not v["name"]:
                skipped += 1
                continue
            existing = by_id.get(v["id"])
            if existing:
                for k, val in v.items():
                    if val and k not in ("id",):
                        existing[k] = val
                updated += 1
            else:
                v.setdefault("discovered_at", now_iso()[:10])
                data.setdefault("vendors", []).append(v)
                by_id[v["id"]] = v
                added += 1
    if not dry_run:
        save_list(data)
    return {
        "ok": True,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "total": len(data.get("vendors") or []),
        "path": str(LIST_PATH),
        "dry_run": dry_run,
    }


def merge_vendors(new_vendors: list[dict[str, Any]], *, dry_run: bool) -> dict[str, Any]:
    data = load_list()
    by_id = vendor_index(data)
    added = updated = 0
    for raw in new_vendors:
        if not isinstance(raw, dict):
            continue
        v = row_to_vendor({k: str(val or "") for k, val in raw.items()}, source="grok_discovery")
        if not v["name"]:
            continue
        existing = by_id.get(v["id"])
        if existing:
            for k, val in v.items():
                if val:
                    existing[k] = val
            updated += 1
        else:
            v.setdefault("discovered_at", now_iso()[:10])
            v.setdefault("status", "discovered")
            data.setdefault("vendors", []).append(v)
            by_id[v["id"]] = v
            added += 1
    if not dry_run:
        save_list(data)
    return {"ok": True, "added": added, "updated": updated, "dry_run": dry_run}


DISCOVERY_BLOCK_MARKER_RE = re.compile(
    r"📎\s*Jarvis\s*用[（(]探索追記[）)]",
    re.IGNORECASE,
)


def parse_discovery_yaml_from_text(text: str) -> list[dict[str, Any]]:
    """部長日報・Grok 出力から vendors: YAML を抽出。"""
    chunks: list[str] = []
    for m in re.finditer(r"```(?:yaml|yml)?\s*\n(.*?)```", text, re.DOTALL | re.IGNORECASE):
        chunks.append(m.group(1))
    marker = DISCOVERY_BLOCK_MARKER_RE.search(text or "")
    if marker:
        tail = text[marker.end() :]
        tail = re.split(r"\n📎", tail, maxsplit=1)[0]
        chunks.append(tail)
    vendors: list[dict[str, Any]] = []
    for raw in chunks:
        block = (raw or "").strip()
        if not block or "vendors:" not in block:
            continue
        try:
            parsed = yaml.safe_load(block)
        except Exception:
            continue
        if isinstance(parsed, dict) and isinstance(parsed.get("vendors"), list):
            vendors.extend(v for v in parsed["vendors"] if isinstance(v, dict))
        elif isinstance(parsed, list):
            vendors.extend(v for v in parsed if isinstance(v, dict))
    return vendors


def merge_vendors_from_text(text: str, *, dry_run: bool) -> dict[str, Any]:
    vendors = parse_discovery_yaml_from_text(text)
    if not vendors:
        return {
            "ok": True,
            "skipped": "no discovery yaml",
            "parsed": 0,
            "added": 0,
            "updated": 0,
            "dry_run": dry_run,
        }
    out = merge_vendors(vendors, dry_run=dry_run)
    out["parsed"] = len(vendors)
    return out


def summary(data: dict[str, Any] | None = None) -> dict[str, Any]:
    data = data or load_list()
    counts: dict[str, int] = {}
    for v in data.get("vendors") or []:
        st = str(v.get("status") or "pending")
        counts[st] = counts.get(st, 0) + 1
    settings = data.get("settings") or {}
    ops_ref = sum(
        1 for v in (data.get("vendors") or []) if v.get("ops_contacted_at")
    )
    return {
        "ok": True,
        "path": str(LIST_PATH),
        "exists": LIST_PATH.is_file(),
        "total": len(data.get("vendors") or []),
        "by_status": counts,
        "ops_contacted_reference": ops_ref,
        "outreach_from": settings.get("outreach_from", "matsuno.estate@gmail.com"),
        "daily_outreach_limit": settings.get("daily_outreach_limit", 3),
        "daily_discovery_limit": settings.get("daily_discovery_limit", 5),
        "target_areas": settings.get("target_areas") or [],
    }


def next_pending(*, limit: int, statuses: tuple[str, ...] = ("pending", "discovered")) -> list[dict[str, Any]]:
    data = load_list()
    out: list[dict[str, Any]] = []
    for v in data.get("vendors") or []:
        st = str(v.get("status") or "pending")
        if st in statuses and v.get("name"):
            out.append(v)
    out.sort(
        key=lambda x: (
            REGION_ORDER.get(x.get("list_region") or "chubu", 9),
            int(x.get("list_no") or 99999),
            x.get("name") or "",
        )
    )
    return out[:limit]


def _domain_key(url: str) -> str:
    s = (url or "").strip()
    if not s:
        return ""
    try:
        host = urlparse(s).netloc.lower()
        if host.startswith("www."):
            host = host[4:]
        return host
    except Exception:
        return ""


def outreach_route_key(v: dict[str, Any]) -> str:
    cu = str(v.get("contact_url") or v.get("url") or "").strip()
    if not cu:
        return ""
    try:
        p = urlparse(cu)
        host = p.netloc.lower()
        if host.startswith("www."):
            host = host[4:]
        path = (p.path or "/").rstrip("/") or "/"
        return f"route:{host}{path}"
    except Exception:
        return ""


def vendor_has_url(v: dict[str, Any]) -> bool:
    return bool(str(v.get("contact_url") or "").strip() or str(v.get("url") or "").strip())


def discovery_query(v: dict[str, Any]) -> str:
    name = str(v.get("name") or "").strip()
    pref = str(v.get("prefecture") or "").strip()
    city = str(v.get("city") or "").strip()
    area = str(v.get("area") or "").strip()
    loc = area or f"{pref}{city}"
    return f"{name} {loc} 不動産 公式 問い合わせ"


def infer_group_key(v: dict[str, Any]) -> str:
    explicit = str(v.get("group_key") or v.get("chain_id") or "").strip()
    if explicit:
        return explicit
    name = str(v.get("name") or "")
    for key, pat in CHAIN_PATTERNS:
        if re.search(pat, name, re.I):
            return key
    contact = str(v.get("contact_url") or v.get("url") or "")
    dom = _domain_key(contact)
    if dom:
        return f"domain:{dom}"
    return f"standalone:{v.get('id') or slug_id(name, str(v.get('area') or ''))}"


def enrich_vendor(v: dict[str, Any]) -> dict[str, Any]:
    out = dict(v)
    out["group_key"] = infer_group_key(v)
    out["outreach_route_key"] = outreach_route_key(v)
    if not vendor_has_url(v):
        out["needs_url_discovery"] = True
        out["discovery_query"] = discovery_query(v)
    else:
        out["needs_url_discovery"] = False
    return out


def load_form_contact() -> dict[str, str] | None:
    if not JARVIS_PRIVATE.is_file():
        return None
    vals: dict[str, str] = {}
    for line in JARVIS_PRIVATE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, rest = line.partition("=")
        vals[k.strip()] = rest.strip().strip('"').strip("'")
    name = vals.get("PERSONAL_NAME", "")
    if not name:
        return None
    family = vals.get("PERSONAL_NAME_KANA_FAMILY", "")
    given = vals.get("PERSONAL_NAME_KANA_GIVEN", "")
    return {
        "name": name,
        "name_kana": f"{family}{given}",
        "email": "matsuno.estate@gmail.com",
        "phone": vals.get("PERSONAL_PHONE", ""),
        "postal_code": vals.get("HOME_POSTAL_CODE", ""),
        "address": vals.get("HOME_ADDRESS", ""),
    }


def outreach_phase_settings(settings: dict[str, Any]) -> tuple[int, int]:
    phase = int(settings.get("outreach_phase") or 1)
    if phase not in PHASE_DAILY_LIMIT:
        phase = 1
    daily = int(settings.get("daily_outreach_limit") or PHASE_DAILY_LIMIT[phase])
    cap = PHASE_DAILY_LIMIT[phase]
    if daily > cap:
        daily = cap
    if daily < 1:
        daily = cap
    return phase, daily


def discovery_daily_limit(settings: dict[str, Any]) -> int:
    phase, _ = outreach_phase_settings(settings)
    phase_cap = 3 if phase == 1 else 5
    limit = int(settings.get("daily_discovery_limit") or phase_cap)
    if limit > phase_cap:
        limit = phase_cap
    if limit < 1:
        limit = phase_cap
    return limit


def contacted_snapshot(data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    contacted: list[dict[str, Any]] = []
    routes: set[str] = set()
    brands: set[str] = set()
    for v in data.get("vendors") or []:
        if not isinstance(v, dict) or not v.get("id"):
            continue
        st = str(v.get("status") or "")
        if st not in {"contacted", "replied"} and not (v.get("contacted_at") or "").strip():
            continue
        gk = infer_group_key(v)
        rk = outreach_route_key(v)
        cu = str(v.get("contact_url") or v.get("url") or "")
        contacted.append(
            {
                "id": str(v["id"]),
                "name": str(v.get("name") or ""),
                "status": st,
                "group_key": gk,
                "outreach_route_key": rk,
                "contact_url": cu,
                "contacted_at": str(v.get("contacted_at") or ""),
            }
        )
        if rk:
            routes.add(rk)
        brands.add(gk)
    contacted.sort(key=lambda x: x.get("id") or "")
    return contacted, sorted(routes), sorted(brands)


def iso_week_id(d: date | None = None) -> str:
    d = d or date.today()
    iso = d.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def build_batch_week_payload(
    *,
    batch_days: int = 7,
    outreach_phase: int | None = None,
) -> dict[str, Any]:
    data = load_list()
    settings = data.get("settings") or {}
    phase, daily_limit = outreach_phase_settings(settings)
    if outreach_phase is not None:
        phase = outreach_phase
        daily_limit = PHASE_DAILY_LIMIT.get(phase, 3)
    queue_size = batch_days * daily_limit
    pending = [enrich_vendor(v) for v in next_pending(limit=queue_size)]
    contacted_vendors, routes_contacted, group_contacted = contacted_snapshot(data)
    today = date.today().isoformat()
    batch_id = iso_week_id()
    form_contact = load_form_contact()
    payload: dict[str, Any] = {
        "ok": True,
        "mode": "batch_week",
        "batch_id": batch_id,
        "batch_days": batch_days,
        "batch_start": today,
        "outreach_phase": phase,
        "daily_limit": daily_limit,
        "weekly_quota": batch_days * daily_limit,
        "outreach_from": settings.get("outreach_from", "matsuno.estate@gmail.com"),
        "outreach_note": settings.get(
            "outreach_note",
            "送信は松野個人・松野エステイト（matsuno.estate）。",
        ),
        "batch_progress": {
            "sent_ids": [],
            "skipped_ids": [],
            "failed_ids": [],
            "last_run_date": None,
            "days_completed": 0,
        },
        "contacted_vendors": contacted_vendors,
        "routes_contacted": routes_contacted,
        "group_contacted": group_contacted,
        "vendors": pending,
    }
    if form_contact:
        payload["form_contact"] = form_contact
    else:
        payload["form_contact_missing"] = True
        payload["form_contact_note"] = (
            "Mac .env.jarvis_private から form_contact を読めませんでした。"
            "送信前に松野へ確認。"
        )
    return payload


def grok_weekly_kickoff(payload: dict[str, Any]) -> str:
    phase = payload.get("outreach_phase")
    daily = payload.get("daily_limit")
    batch_id = payload.get("batch_id")
    n = len(payload.get("vendors") or [])
    contacted_n = len(payload.get("contacted_vendors") or [])
    first_id = (payload.get("vendors") or [{}])[0].get("id", "—")
    return f"""【週次バッチ開始 · Phase {phase} · {batch_id}】

添付 JSON が今週の作業キュー（最大 {n} 社 · 1日 {daily} 社上限 · {payload.get('batch_days')} 日分）。
同じスレッド内で batch_progress を保持し、毎日「本日分」と言ったらその日の {daily} 社だけ送信してください（土日も同じ）。

ルール:
- 1日 {daily} 社まで（Phase {phase}）。超えない
- 系列重複は skip（Instructions どおり）。スキップ分を別店で埋めない
- 送信ごとに --mark 行を報告
- その日終了時: 「本日完了（成功X / skipY / 失敗Z）」+ 日次 [Grok部長] 日報メール
- 土曜 or 日曜: 週次 [Grok部長] 週次 メール + 全 --mark 行一覧再掲
- キュー残り0または不足: 松野へ Jarvis で --batch-week --grok-kickoff 再生成を依頼（Instructions §週次バッチ）

既 contact 済み: {contacted_n} 社（JSON の contacted_vendors 参照）。
キュー先頭: {first_id} から。

以下 JSON です。"""


def extract_discovered_url(note: str) -> str:
    m = DISCOVERED_URL_RE.search(note or "")
    if not m:
        return ""
    return m.group(1).rstrip(".,)")


def parse_mark_commands(text: str) -> list[dict[str, str]]:
    """Grok サマリー等から --mark 行を抽出（同一 id は最後 wins）。"""
    by_id: dict[str, dict[str, str]] = {}
    for line in text.splitlines():
        for m in MARK_CMD_RE.finditer(line):
            note = (m.group("note_dq") or m.group("note_sq") or "").strip()
            by_id[m.group("id")] = {
                "id": m.group("id"),
                "status": m.group("status").lower(),
                "note": note,
            }
    return list(by_id.values())


def apply_marks_from_text(text: str, *, dry_run: bool) -> dict[str, Any]:
    cmds = parse_mark_commands(text)
    if not cmds:
        return {"ok": False, "error": "no --mark commands found", "parsed": 0}
    data = load_list()
    by_id = vendor_index(data)
    applied: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    url_updates: list[dict[str, str]] = []
    today = now_iso()[:10]
    for cmd in cmds:
        vid = cmd["id"]
        status = cmd["status"]
        note = cmd["note"]
        if status not in STATUSES:
            errors.append({"id": vid, "error": f"invalid status: {status}"})
            continue
        v = by_id.get(vid)
        if not v:
            errors.append({"id": vid, "error": "vendor not found"})
            continue
        v["status"] = status
        if note:
            v["notes"] = note if not v.get("notes") else f"{v['notes']} | {note}"
        if status == "contacted":
            v["contacted_at"] = today
        if status == "replied":
            v["replied_at"] = today
        discovered = extract_discovered_url(note)
        if discovered:
            if not str(v.get("contact_url") or "").strip():
                v["contact_url"] = discovered
                url_updates.append({"id": vid, "contact_url": discovered})
            if not str(v.get("url") or "").strip():
                v["url"] = discovered
        v["updated_at"] = now_iso()
        applied.append({"id": vid, "status": status, "note": note})
    if not dry_run and applied:
        save_list(data)
    return {
        "ok": len(errors) == 0,
        "parsed": len(cmds),
        "applied": len(applied),
        "applied_items": applied,
        "url_updates": url_updates,
        "errors": errors,
        "dry_run": dry_run,
    }


def mark_vendor(
    vid: str,
    *,
    status: str,
    note: str = "",
    result: str = "",
    dry_run: bool,
) -> dict[str, Any]:
    if status not in STATUSES:
        return {"ok": False, "error": f"invalid status: {status}"}
    data = load_list()
    by_id = vendor_index(data)
    v = by_id.get(vid)
    if not v:
        return {"ok": False, "error": f"vendor not found: {vid}"}
    today = now_iso()[:10]
    v["status"] = status
    if note:
        v["notes"] = note if not v.get("notes") else f"{v['notes']} | {note}"
    if result:
        v["last_result"] = result
    if status == "contacted":
        v["contacted_at"] = today
        if not note:
            note = "個人送信(estate)"
    if status == "replied":
        v["replied_at"] = today
    v["updated_at"] = now_iso()
    if not dry_run:
        save_list(data)
    return {"ok": True, "vendor": v, "dry_run": dry_run}


def grok_discovery_prompt() -> str:
    data = load_list()
    settings = data.get("settings") or {}
    limit = discovery_daily_limit(settings)
    phase, _ = outreach_phase_settings(settings)
    areas = settings.get("target_areas") or ["愛知県", "岐阜県", "三重県"]
    existing = [
        f"- {v.get('name')} ({v.get('area') or '—'})"
        for v in (data.get("vendors") or [])[:40]
    ]
    areas_s = "、".join(areas)
    block = "\n".join(existing) if existing else "（まだなし）"
    return f"""【Grok 業者探索 — 日次 {limit} 件 · Phase {phase}】

対象: {areas_s} の **戸建・投資向け地場不動産**（大手全国チェーンは優先度低）
1日 **{limit} 社まで** 新規発見し、下記 YAML 追記ブロックで出力すること。
既存リストと **重複会社は載せない**。

## 既存（抜粋・重複禁止）
{block}

## 各社の調査
- 公式サイト・問合せURL
- 戸建/投資実績の有無（なければ skip 候補）
- 問合せは **まだ送らない**（discovered のみ。送信は別タスク）

## 出力形式（このブロックだけ返す）
```yaml
vendors:
  - name: "会社名"
    area: "愛知県〇〇市"
    prefecture: "愛知県"
    city: "〇〇市"
    url: "https://..."
    contact_url: "https://.../contact"
    channel: web_form
    contact_email: ""
    status: discovered
    source: grok_discovery
    notes: "戸建あり。理由1行"
```

完了後: 「追加 N 件」と1行サマリー。
"""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--summary", action="store_true")
    ap.add_argument("--import-csv", metavar="PATH")
    ap.add_argument(
        "--import-xlsx",
        metavar="PATH",
        help="神大家地場リスト Excel（地場不動産業者一覧_*.xlsx）",
    )
    ap.add_argument("--merge-append", metavar="PATH", help="Grok 追記 YAML をマージ")
    ap.add_argument("--next", type=int, default=0, help="次に問合せする pending 件数表示")
    ap.add_argument(
        "--batch-week",
        action="store_true",
        help="Grok 週次バッチ用 JSON（batch_days × daily_limit 件）",
    )
    ap.add_argument(
        "--batch-days",
        type=int,
        default=7,
        help="--batch-week の日数（既定7）",
    )
    ap.add_argument(
        "--outreach-phase",
        type=int,
        default=0,
        help="Phase 上書き（1|2|3）。0=YAML settings 従う",
    )
    ap.add_argument(
        "--grok-kickoff",
        action="store_true",
        help="--batch-week 時に Grok キックオフ文を JSON の前に出力",
    )
    ap.add_argument(
        "--kickoff-only",
        action="store_true",
        help="--batch-week と併用。キックオフ文のみ（JSON なし）",
    )
    ap.add_argument("--mark", metavar="ID")
    ap.add_argument("--status", default="contacted")
    ap.add_argument("--note", default="")
    ap.add_argument("--result", default="")
    ap.add_argument("--grok-discovery-prompt", action="store_true")
    ap.add_argument(
        "--apply-marks",
        metavar="PATH",
        help="Grok サマリー等の --mark 行を一括反映（- で stdin）",
    )
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--merge",
        action="store_true",
        help="--import-xlsx 時に既存リストを残して追加・更新",
    )
    args = ap.parse_args()

    if args.grok_discovery_prompt:
        print(grok_discovery_prompt())
        return 0

    if args.import_xlsx:
        out = import_kamiooya_xlsx(
            Path(args.import_xlsx), dry_run=args.dry_run, merge=args.merge
        )
        print(json.dumps(out, ensure_ascii=False, indent=2))
        print(f"VENDOR_LIST_RESULT:{json.dumps(out, ensure_ascii=False)}")
        return 0

    if args.import_csv:
        out = import_csv(Path(args.import_csv), dry_run=args.dry_run)
        print(json.dumps(out, ensure_ascii=False, indent=2))
        print(f"VENDOR_LIST_RESULT:{json.dumps(out, ensure_ascii=False)}")
        return 0

    if args.merge_append:
        raw = yaml.safe_load(Path(args.merge_append).read_text(encoding="utf-8")) or {}
        vendors = raw.get("vendors") if isinstance(raw, dict) else raw
        if not isinstance(vendors, list):
            print(json.dumps({"ok": False, "error": "no vendors list in file"}, ensure_ascii=False))
            return 1
        out = merge_vendors(vendors, dry_run=args.dry_run)
        print(json.dumps(out, ensure_ascii=False, indent=2))
        print(f"VENDOR_LIST_RESULT:{json.dumps(out, ensure_ascii=False)}")
        return 0

    if args.mark:
        out = mark_vendor(
            args.mark,
            status=args.status,
            note=args.note,
            result=args.result,
            dry_run=args.dry_run,
        )
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 0 if out.get("ok") else 1

    if args.apply_marks:
        p = args.apply_marks
        if p == "-":
            text = sys.stdin.read()
        else:
            text = Path(p).read_text(encoding="utf-8")
        out = apply_marks_from_text(text, dry_run=args.dry_run)
        print(json.dumps(out, ensure_ascii=False, indent=2))
        print(f"VENDOR_LIST_RESULT:{json.dumps(out, ensure_ascii=False)}")
        return 0 if out.get("ok") else 1

    if args.next > 0:
        items = next_pending(limit=args.next)
        data = load_list()
        settings = data.get("settings") or {}
        phase, daily_limit = outreach_phase_settings(settings)
        contacted_vendors, routes_contacted, group_contacted = contacted_snapshot(data)
        body: dict[str, Any] = {
            "ok": True,
            "count": len(items),
            "outreach_phase": phase,
            "daily_limit": daily_limit,
            "outreach_from": settings.get(
                "outreach_from", "matsuno.estate@gmail.com"
            ),
            "outreach_note": settings.get("outreach_note", ""),
            "contacted_vendors": contacted_vendors,
            "routes_contacted": routes_contacted,
            "group_contacted": group_contacted,
            "vendors": [enrich_vendor(v) for v in items],
        }
        fc = load_form_contact()
        if fc:
            body["form_contact"] = fc
        print(json.dumps(body, ensure_ascii=False, indent=2))
        return 0

    if args.batch_week:
        phase_override = args.outreach_phase if args.outreach_phase > 0 else None
        payload = build_batch_week_payload(
            batch_days=max(1, args.batch_days),
            outreach_phase=phase_override,
        )
        if args.grok_kickoff or args.kickoff_only:
            print(grok_weekly_kickoff(payload))
            print()
        if not args.kickoff_only:
            print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0

    if args.summary or len(sys.argv) == 1:
        out = summary()
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 0

    ap.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
