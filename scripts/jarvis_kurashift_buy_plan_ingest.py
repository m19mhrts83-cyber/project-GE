#!/usr/bin/env python3
"""KURASHIFT: 買い進めプラン Excel を Supabase へ取込。

  cd ~/git-repos && set -a && source .env.jarvis_private && set +a
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_buy_plan_ingest.py
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_buy_plan_ingest.py --dry-run
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_buy_plan_ingest.py --canonical-only
"""
from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
BUY_DIR = Path(
    "/Users/matsunomasaharu2/Library/CloudStorage/OneDrive-個人用/"
    "215_神・大家さん倶楽部/05_【購入】買い進めプランニング"
)
CANONICAL_KEY = "251124"
STEP3_NAME = "STEP3 夢を叶えるプランニングシート（ver3.0)"
AREA_NAME = "物件購入検討エリア・条件"
CONSTRAINT_NAME = "プランニング制約"
KIMURA_NAME = "木村さんコメント"
EXCEL_EPOCH = datetime(1899, 12, 30)


def jst_now() -> datetime:
    return datetime.now(timezone.utc).astimezone()


def sb_client() -> Any:
    url = os.environ.get("JARVIS_SUPABASE_URL")
    key = os.environ.get("JARVIS_SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("JARVIS_SUPABASE_URL / JARVIS_SUPABASE_SERVICE_ROLE_KEY が必要です")
    from supabase import create_client

    return create_client(url, key)


def file_checksum(path: Path, limit: int = 2_000_000) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        remaining = limit
        while remaining > 0:
            chunk = f.read(min(65536, remaining))
            if not chunk:
                break
            h.update(chunk)
            remaining -= len(chunk)
    return h.hexdigest()


def version_key_from_name(name: str) -> str | None:
    m = re.search(r"(\d{6})", name)
    return m.group(1) if m else None


def as_of_from_key(key: str) -> date:
    # YYMMDD → 20YY-MM-DD
    y, m, d = int(key[:2]), int(key[2:4]), int(key[4:6])
    return date(2000 + y, m, d)


def excel_to_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return (EXCEL_EPOCH + timedelta(days=float(val))).date()
        except Exception:
            return None
    return None


def cell_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).replace("\n", " ").strip()
    return s or None


def cell_num(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("%", "").strip())
    except Exception:
        return None


def list_xlsx(*, canonical_only: bool) -> list[Path]:
    out: list[Path] = []
    for p in sorted(BUY_DIR.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        key = version_key_from_name(p.name)
        if not key:
            continue
        if canonical_only and key != CANONICAL_KEY:
            continue
        out.append(p)
    return out


def parse_step3(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in range(9, (ws.max_row or 9) + 1):
        action = cell_str(ws.cell(r, 3).value)
        if not action:
            continue
        if action not in ("購入", "売却", "資金調達") and "購" not in action and "売" not in action:
            # 空行・説明行スキップ
            if not ws.cell(r, 2).value and not ws.cell(r, 7).value:
                continue
        no = ws.cell(r, 2).value
        row_no = int(no) if isinstance(no, (int, float)) else r
        price = cell_num(ws.cell(r, 10).value)
        yld = cell_num(ws.cell(r, 11).value)
        if yld is not None and yld > 1:
            yld = yld / 100.0
        rate = cell_num(ws.cell(r, 19).value)
        if rate is not None and rate > 1:
            rate = rate / 100.0
        rows.append(
            {
                "row_no": row_no,
                "action": action,
                "entity": cell_str(ws.cell(r, 6).value),
                "location": cell_str(ws.cell(r, 7).value),
                "structure": cell_str(ws.cell(r, 8).value),
                "built_year": cell_str(ws.cell(r, 9).value),
                "price_man": price,
                "yield_pct": yld,
                "reno_man": cell_num(ws.cell(r, 12).value),
                "cost_man": cell_num(ws.cell(r, 14).value),
                "bank": cell_str(ws.cell(r, 16).value),
                "loan_man": cell_num(ws.cell(r, 17).value),
                "down_man": cell_num(ws.cell(r, 18).value),
                "rate_pct": rate,
                "loan_years": cell_num(ws.cell(r, 20).value),
                "property_name": cell_str(ws.cell(r, 38).value),  # AL
                "sale_strategy": cell_str(ws.cell(r, 40).value),  # AN
                "event_date": excel_to_date(ws.cell(r, 5).value),
                "memo": cell_str(ws.cell(r, 1).value),
                "extras": {
                    "kind_label": cell_str(ws.cell(r, 15).value),
                    "self_funds": cell_num(ws.cell(r, 4).value),
                    "yield_after_reno": cell_num(ws.cell(r, 13).value),
                },
            }
        )
    return rows


def parse_area(ws) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    kind = "area"
    for r in range(1, (ws.max_row or 1) + 1):
        raw = cell_str(ws.cell(r, 1).value)
        if not raw:
            continue
        if "購入希望" in raw:
            kind = "purchase_rule"
        elif "希望エリア" in raw:
            kind = "area"
        out.append(
            {
                "kind": kind if kind in ("area", "purchase_rule") else "other",
                "sort_order": r,
                "payload": {"line": raw},
                "raw_text": raw,
            }
        )
    return out


def parse_constraints(ws) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in range(2, (ws.max_row or 2) + 1):
        lender = cell_str(ws.cell(r, 4).value)
        collateral = cell_str(ws.cell(r, 2).value)
        if not lender and not collateral:
            continue
        out.append(
            {
                "row_no": r,
                "updated_on": excel_to_date(ws.cell(r, 1).value),
                "collateral_type": collateral,
                "guarantor": cell_str(ws.cell(r, 3).value),
                "lender": lender,
                "limit_note": cell_str(ws.cell(r, 5).value),
                "attr_note": cell_str(ws.cell(r, 6).value),
                "rate_term": cell_str(ws.cell(r, 7).value),
                "prop_cond": cell_str(ws.cell(r, 8).value),
                "income_cond": cell_str(ws.cell(r, 9).value),
                "geo_cond": cell_str(ws.cell(r, 10).value),
                "extras": {"entity": cell_str(ws.cell(r, 11).value)},
            }
        )
    return out


def parse_notes(ws, sheet_name: str) -> dict[str, Any]:
    lines: list[str] = []
    for r in range(1, min((ws.max_row or 1), 200) + 1):
        vals = [cell_str(ws.cell(r, c).value) for c in range(1, 6)]
        vals = [v for v in vals if v]
        if vals:
            lines.append(" | ".join(vals))
    return {"sheet_name": sheet_name, "payload": {"lines": lines, "line_count": len(lines)}}


def upsert_version(
    sb: Any,
    path: Path,
    *,
    dry_run: bool,
) -> dict[str, Any]:
    key = version_key_from_name(path.name)
    assert key
    st = path.stat()
    as_of = as_of_from_key(key)
    is_canonical = key == CANONICAL_KEY
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    # read_only だと一部操作が制限されるため通常モードで再読込
    wb.close()
    wb = load_workbook(path, data_only=True)

    events: list[dict[str, Any]] = []
    criteria: list[dict[str, Any]] = []
    constraints: list[dict[str, Any]] = []
    notes: list[dict[str, Any]] = []
    notes_status = "ok"

    if STEP3_NAME in wb.sheetnames:
        events = parse_step3(wb[STEP3_NAME])
    else:
        notes_status = "partial"
    if AREA_NAME in wb.sheetnames:
        criteria = parse_area(wb[AREA_NAME])
    if CONSTRAINT_NAME in wb.sheetnames:
        constraints = parse_constraints(wb[CONSTRAINT_NAME])
    for sn in wb.sheetnames:
        if "木村" in sn or sn == KIMURA_NAME:
            notes.append(parse_notes(wb[sn], sn))

    meta = {
        "checksum": file_checksum(path),
        "event_count": len(events),
        "criteria_count": len(criteria),
        "constraint_count": len(constraints),
        "ingested_at": jst_now().isoformat(),
    }
    version_row = {
        "version_key": key,
        "as_of": as_of.isoformat(),
        "source_path": str(path),
        "source_filename": path.name,
        "label": path.stem,
        "is_canonical": is_canonical,
        "file_mtime": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
        "file_size_bytes": st.st_size,
        "extract_status": notes_status if events or criteria else "failed",
        "extract_notes": None if events else "STEP3 rows empty or sheet missing",
        "sheet_names": sheet_names,
        "metadata": meta,
        "updated_at": jst_now().isoformat(),
    }
    summary = {
        "key": key,
        "events": len(events),
        "criteria": len(criteria),
        "constraints": len(constraints),
        "notes": len(notes),
        "canonical": is_canonical,
    }
    if dry_run:
        print(f"# dry-run {path.name}: {summary}")
        return summary

    if is_canonical:
        sb.table("kurashift_buy_plan_versions").update({"is_canonical": False}).eq(
            "is_canonical", True
        ).execute()

    up = (
        sb.table("kurashift_buy_plan_versions")
        .upsert(version_row, on_conflict="version_key")
        .execute()
    )
    vid = (up.data or [{}])[0].get("id")
    if not vid:
        got = (
            sb.table("kurashift_buy_plan_versions")
            .select("id")
            .eq("version_key", key)
            .limit(1)
            .execute()
        )
        vid = (got.data or [{}])[0].get("id")
    if not vid:
        raise RuntimeError(f"version id missing for {key}")

    sb.table("kurashift_buy_plan_events").delete().eq("version_id", vid).execute()
    sb.table("kurashift_buy_plan_criteria").delete().eq("version_id", vid).execute()
    sb.table("kurashift_buy_plan_constraints").delete().eq("version_id", vid).execute()
    sb.table("kurashift_buy_plan_notes").delete().eq("version_id", vid).execute()

    def chunks(items: list, n: int = 200):
        for i in range(0, len(items), n):
            yield items[i : i + n]

    for batch in chunks([{**e, "version_id": vid, "event_date": e["event_date"].isoformat() if e.get("event_date") else None} for e in events]):
        sb.table("kurashift_buy_plan_events").insert(batch).execute()
    for batch in chunks([{**c, "version_id": vid} for c in criteria]):
        sb.table("kurashift_buy_plan_criteria").insert(batch).execute()
    for batch in chunks(
        [
            {
                **c,
                "version_id": vid,
                "updated_on": c["updated_on"].isoformat() if c.get("updated_on") else None,
            }
            for c in constraints
        ]
    ):
        sb.table("kurashift_buy_plan_constraints").insert(batch).execute()
    for n in notes:
        sb.table("kurashift_buy_plan_notes").upsert(
            {**n, "version_id": vid}, on_conflict="version_id,sheet_name"
        ).execute()

    print(f"# ok {path.name}: {summary} version_id={vid}")
    return summary


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--canonical-only", action="store_true")
    args = ap.parse_args()
    if not BUY_DIR.is_dir():
        raise SystemExit(f"買い進めフォルダが無い: {BUY_DIR}")
    files = list_xlsx(canonical_only=args.canonical_only)
    if not files:
        raise SystemExit("対象 xlsx なし")
    sb = None if args.dry_run else sb_client()
    totals = {"files": 0, "events": 0}
    for p in files:
        s = upsert_version(sb, p, dry_run=args.dry_run)
        totals["files"] += 1
        totals["events"] += int(s.get("events") or 0)
    print(f"📎 buy_plan_ingest: files={totals['files']} events={totals['events']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
