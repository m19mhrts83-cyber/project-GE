#!/usr/bin/env python3
"""canonical 買い進めプランを STEP3 互換の簡易 Excel に export（運営共有用）。

  cd ~/git-repos && set -a && source .env.jarvis_private && set +a
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_buy_plan_export.py
"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[1]
OUT_DIR = Path(
    "/Users/matsunomasaharu2/Library/CloudStorage/OneDrive-個人用/"
    "215_神・大家さん倶楽部/05_【購入】買い進めプランニング/exports"
)
STEP3 = "STEP3 夢を叶えるプランニングシート（ver3.0)"


def sb_client() -> Any:
    url = os.environ.get("JARVIS_SUPABASE_URL")
    key = os.environ.get("JARVIS_SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("JARVIS_SUPABASE_URL / JARVIS_SUPABASE_SERVICE_ROLE_KEY が必要です")
    from supabase import create_client

    return create_client(url, key)


HEADERS = [
    "メモ",
    "No.",
    "購入・売却",
    "自己資金",
    "時期",
    "個人/法人",
    "立地",
    "構造",
    "築年",
    "価格(万円)",
    "利回り",
    "リフォーム額",
    "諸経費",
    "銀行",
    "融資額",
    "頭金",
    "金利",
    "借入年数",
    "物件名",
    "売却戦略",
]


def main() -> int:
    sb = sb_client()
    ver = (
        sb.table("kurashift_buy_plan_versions")
        .select("id, version_key, label, as_of")
        .eq("is_canonical", True)
        .limit(1)
        .execute()
    )
    v = (ver.data or [None])[0]
    if not v:
        raise SystemExit("canonical buy plan がありません。先に ingest してください")
    events = (
        sb.table("kurashift_buy_plan_events")
        .select("*")
        .eq("version_id", v["id"])
        .order("row_no")
        .execute()
    )
    criteria = (
        sb.table("kurashift_buy_plan_criteria")
        .select("kind, raw_text, sort_order")
        .eq("version_id", v["id"])
        .order("sort_order")
        .execute()
    )
    constraints = (
        sb.table("kurashift_buy_plan_constraints")
        .select("*")
        .eq("version_id", v["id"])
        .order("row_no")
        .execute()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = STEP3[:31]
    # 簡易ヘッダ（運営共有用の互換骨格。完全レイアウトではない）
    ws.append([f"export from KURASHIFT canonical={v['version_key']} as_of={v.get('as_of')}"])
    ws.append([])
    ws.append(HEADERS)
    for e in events.data or []:
        ws.append(
            [
                e.get("memo"),
                e.get("row_no"),
                e.get("action"),
                (e.get("extras") or {}).get("self_funds"),
                e.get("event_date"),
                e.get("entity"),
                e.get("location"),
                e.get("structure"),
                e.get("built_year"),
                e.get("price_man"),
                e.get("yield_pct"),
                e.get("reno_man"),
                e.get("cost_man"),
                e.get("bank"),
                e.get("loan_man"),
                e.get("down_man"),
                e.get("rate_pct"),
                e.get("loan_years"),
                e.get("property_name"),
                e.get("sale_strategy"),
            ]
        )

    ws2 = wb.create_sheet("物件購入検討エリア・条件")
    for c in criteria.data or []:
        ws2.append([c.get("raw_text")])

    ws3 = wb.create_sheet("プランニング制約")
    ws3.append(
        [
            "更新日",
            "種別",
            "保証",
            "融資先",
            "融資枠",
            "個人属性",
            "期間/金利",
            "条件_物件",
            "条件_収入",
            "条件_地理",
        ]
    )
    for c in constraints.data or []:
        ws3.append(
            [
                c.get("updated_on"),
                c.get("collateral_type"),
                c.get("guarantor"),
                c.get("lender"),
                c.get("limit_note"),
                c.get("attr_note"),
                c.get("rate_term"),
                c.get("prop_cond"),
                c.get("income_cond"),
                c.get("geo_cond"),
            ]
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out = OUT_DIR / f"KURASHIFT_STEP3export_{v['version_key']}_{stamp}.xlsx"
    wb.save(out)
    print(f"📎 buy_plan_export: {out}")
    print(f"  events={len(events.data or [])} criteria={len(criteria.data or [])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
