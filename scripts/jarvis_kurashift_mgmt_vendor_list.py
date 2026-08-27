#!/usr/bin/env python3
"""S9 管理会社開拓リスト — Excel 取込・状態・生存確認。

正本 YAML: config/kurashift_mgmt_vendor_list.yaml（.gitignore）
Excel 正本（空室一括送信）は壊さない。ここは KURASHIFT 投影用の副本。

  cd ~/git-repos
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_mgmt_vendor_list.py --summary
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_mgmt_vendor_list.py \\
    --import-xlsx ".../★管理会社一覧.xlsx" --merge
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_mgmt_vendor_list.py --next 2
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_mgmt_vendor_list.py \\
    --mark ID --status contacted --note "Web送信"
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_mgmt_vendor_list.py \\
    --mark-alive ID --alive-status ok --alive-method phone
  ~/selenium_env/venv/bin/python scripts/jarvis_kurashift_mgmt_vendor_list.py --alive-queue --limit 2
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))

from jarvis_vendor_alive_lib import (  # noqa: E402
    ALIVE_STATUSES,
    build_alive_queue,
    effective_alive_status,
    ensure_alive_fields,
    is_alive_ok,
    mark_alive as mark_alive_fields,
)

LIST_PATH = REPO / "config" / "kurashift_mgmt_vendor_list.yaml"
EXAMPLE_PATH = REPO / "config" / "kurashift_mgmt_vendor_list.example.yaml"
SHARE_CFG_PATH = REPO / "config" / "kurashift_mgmt_share_folders.yaml"

DEFAULT_XLSX = Path(
    "/Users/matsunomasaharu2/Library/CloudStorage/OneDrive-個人用/"
    "215_神・大家さん倶楽部/20_【空室対策】【修繕】【売却】/"
    "21_【空室対策】募集,ステージング,物件管理/★管理会社一覧.xlsx"
)

STATUSES = frozenset(
    {"pending", "contacted", "replied", "skip", "discovered", "invalid"}
)
LANES = frozenset({"kita_shiga", "midori_caramel", "both"})
AREA_PHRASE = {
    "kita_shiga": "名古屋市北区（志賀本通駅周辺）",
    "midori_caramel": "名古屋市緑区（キャラメル）",
    "both": "名古屋市北区（志賀本通）および緑区",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_share_urls() -> dict[str, str]:
    """MGMT_SHARE_URL_* または kurashift_mgmt_share_folders.yaml。"""
    import os

    out = {
        "common": (os.environ.get("MGMT_SHARE_URL_COMMON") or "").strip(),
        "kita_shiga": (os.environ.get("MGMT_SHARE_URL_KITA") or "").strip(),
        "midori_caramel": (os.environ.get("MGMT_SHARE_URL_MIDORI") or "").strip(),
    }
    if SHARE_CFG_PATH.is_file():
        try:
            cfg = yaml.safe_load(SHARE_CFG_PATH.read_text(encoding="utf-8")) or {}
            defaults = cfg.get("template_defaults") or {}
            for k in ("common", "kita_shiga", "midori_caramel"):
                if not out[k] and defaults.get(k):
                    out[k] = str(defaults[k]).strip()
        except Exception:
            pass
    return out


def infer_property_lane(
    *,
    property_area: str = "",
    city: str = "",
    notes: str = "",
    explicit: str = "",
) -> str:
    ex = (explicit or "").strip()
    if ex in LANES:
        return ex
    # Excel レーン列の日本語
    if ex in ("北区", "志賀", "kita"):
        return "kita_shiga"
    if ex in ("緑区", "キャラメル", "midori"):
        return "midori_caramel"
    blob = f"{property_area} {city} {notes}"
    has_midori = "緑区" in blob or "キャラメル" in blob
    has_kita = "北区" in blob or "志賀" in blob
    if has_midori and has_kita:
        return "both"
    if has_midori:
        return "midori_caramel"
    return "kita_shiga"


def ensure_precheck_fields(v: dict[str, Any]) -> None:
    if "vacancy_listing_ok" not in v:
        v["vacancy_listing_ok"] = None
    if not v.get("property_lane"):
        v["property_lane"] = infer_property_lane(
            property_area=str(v.get("property_area") or ""),
            city=str(v.get("city") or ""),
            notes=str(v.get("notes") or ""),
        )
    v.setdefault("precheck_sent_at", "")


def slug_id(name: str, area: str = "") -> str:
    base = f"mgmt-{name}-{area}".strip().lower()
    base = re.sub(r"\s+", "-", base)
    base = re.sub(r"[^a-z0-9\-ぁ-んァ-ヶ一-龥]", "", base)
    if len(base) >= 8:
        return base[:48]
    h = hashlib.sha256(base.encode()).hexdigest()[:8]
    return f"mgmt-{h}"


def load_list(path: Path | None = None) -> dict[str, Any]:
    p = path or LIST_PATH
    if not p.is_file():
        if EXAMPLE_PATH.is_file():
            data = yaml.safe_load(EXAMPLE_PATH.read_text(encoding="utf-8")) or {}
            data.setdefault("settings", {})
            data.setdefault("vendors", [])
            return data
        return {"settings": {}, "vendors": []}
    return yaml.safe_load(p.read_text(encoding="utf-8")) or {
        "settings": {},
        "vendors": [],
    }


def save_list(data: dict[str, Any], path: Path | None = None) -> None:
    p = path or LIST_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def vendor_index(data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for v in data.get("vendors") or []:
        if isinstance(v, dict) and v.get("id"):
            out[str(v["id"])] = v
    return out


def _cell(row: tuple[Any, ...], idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip().replace("\n", " ")


def _mark_true(val: str) -> bool:
    s = (val or "").strip()
    return s in ("〇", "○", "有り", "あり", "有", "OK", "ok", "1", "True", "true")


def _status_hint_from_notes(notes: str) -> str:
    n = notes or ""
    if "戸別管理不可" in n or "戸建て管理不可" in n:
        return "skip"
    if "問い合わせ中→OK" in n or "戸別管理問い合わせ中→OK" in n or "管理OK" in n:
        return "replied"
    if "問い合わせ中" in n:
        return "contacted"
    return "pending"


def _status_from_excel_contact(
    *,
    notes: str,
    method: str,
    date_cell: str,
    flyer: str,
    photo: str,
    reply: str,
    result: str,
    memo: str,
) -> str:
    """備考＋方法/日付/チラシ送付/回答など。既アプローチは pending にしない。"""
    status = _status_hint_from_notes(notes)
    if "不可" in notes or "不可" in result:
        return "skip"
    if status in ("replied", "skip"):
        return status
    if any(k in result for k in ("一般仲介", "募集する", "反響", "OK", "可")) and "不可" not in result:
        return "replied"
    if any(str(x or "").strip() for x in (method, date_cell, flyer, photo, reply, result, memo)):
        return "contacted"
    return status


def import_mgmt_xlsx(
    path: Path,
    *,
    dry_run: bool,
    merge: bool = True,
) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl required for --import-xlsx") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # G2 が運用正本（方法・日付・チラシ送付あり）
    if "G2" in wb.sheetnames:
        sheet_name = "G2"
    elif "一覧" in wb.sheetnames:
        sheet_name = "一覧"
    else:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # header row: find row with 会社名
    header_i = 0
    for i, r in enumerate(rows[:10]):
        cells = [str(c or "") for c in (r or ())]
        if any("会社名" in c for c in cells):
            header_i = i
            break
    header = rows[header_i] if rows else ()
    col = {str(h or "").replace("\n", ""): i for i, h in enumerate(header)}

    def idx(*names: str) -> int:
        for n in names:
            if n in col:
                return col[n]
            for k, i in col.items():
                if n in k:
                    return i
        return -1

    i_no = idx("No.")
    i_pref = idx("県")
    i_city = idx("市")
    i_sta = idx("駅")
    i_name = idx("会社名")
    i_area = idx("物件エリア")
    i_url = idx("URL")
    i_notes = idx("備考")
    i_buy = idx("購入")
    i_lease = idx("賃貸仲介")
    i_mgmt = idx("賃貸管理")
    i_repair = idx("修繕")
    i_sale = idx("売却査定")
    i_kodate_ok = idx("戸建管理OK", "戸別管理")
    i_vacancy_ok = idx("空室メール可")
    i_lane = idx("物件レーン", "レーン")
    i_form = idx("問い合わせフォーム")
    i_mail = idx("mail", "個別メール")
    i_tel = idx("tel")
    i_result = idx("問い合わせ結果")
    i_method = idx("方法")
    i_date = idx("日付")
    i_flyer = idx("募集チラシ")
    i_photo = idx("部屋写真")
    i_reply = idx("回答有無")
    i_memo = idx("登録作業状況", "メモ欄")

    existing = load_list() if merge else {"settings": {}, "vendors": []}
    by_id = vendor_index(existing)
    by_name = {
        str(v.get("name") or "").strip(): v
        for v in (existing.get("vendors") or [])
        if isinstance(v, dict) and v.get("name")
    }
    added = 0
    updated = 0
    skipped = 0

    for r in rows[header_i + 1 :]:
        if not r:
            continue
        name = _cell(r, i_name)
        if not name or name in ("記入例", "会社名", "〇〇"):
            continue
        # skip title-ish / placeholder
        if name == "管理会社一覧" or "〇〇" in name:
            continue
        url = _cell(r, i_url)
        if "〇〇" in url or url in ("〇〇",):
            continue
        pref = _cell(r, i_pref)
        city = _cell(r, i_city)
        station = _cell(r, i_sta)
        prop_area = _cell(r, i_area)
        area = " ".join(x for x in (pref, city, station) if x).strip()
        url = _cell(r, i_url)
        if url and not url.startswith("http"):
            url = f"https://{url}"
        notes = _cell(r, i_notes)
        result = _cell(r, i_result)
        method = _cell(r, i_method)
        date_raw = r[i_date] if i_date >= 0 and i_date < len(r) else None
        date_cell = str(date_raw)[:10] if date_raw not in (None, "") else ""
        flyer = _cell(r, i_flyer)
        photo = _cell(r, i_photo)
        reply = _cell(r, i_reply)
        memo = _cell(r, i_memo)
        services = {
            "buy": _mark_true(_cell(r, i_buy)),
            "lease_broker": _mark_true(_cell(r, i_lease)),
            "lease_mgmt": _mark_true(_cell(r, i_mgmt)),
            "repair": _mark_true(_cell(r, i_repair)),
            "sale_appraisal": _mark_true(_cell(r, i_sale)),
            "kodate_mgmt_ok": _mark_true(_cell(r, i_kodate_ok)),
        }
        form = _cell(r, i_form)
        contact_url = url if form in ("有り", "あり", "〇", "○") else url
        email = _cell(r, i_mail)
        phone = _cell(r, i_tel)
        no = _cell(r, i_no)
        vid = slug_id(name, city or pref)
        if no.isdigit():
            vid = f"mgmt-{int(no):03d}-{slug_id(name, '')[5:20]}"

        status = _status_from_excel_contact(
            notes=notes,
            method=method,
            date_cell=date_cell,
            flyer=flyer,
            photo=photo,
            reply=reply,
            result=result,
            memo=memo,
        )

        vac_cell = _cell(r, i_vacancy_ok)
        vacancy_ok: bool | None
        if vac_cell in ("不可", "NG", "ng", "×", "x", "FALSE", "false", "0"):
            vacancy_ok = False
        elif _mark_true(vac_cell) or vac_cell in ("可", "OK"):
            vacancy_ok = True
        else:
            vacancy_ok = None
            if status == "replied" and "不可" not in notes and "不可" not in result:
                vacancy_ok = True
            elif status == "skip":
                vacancy_ok = False

        contact_bits = []
        if method:
            contact_bits.append(f"方法:{method}")
        if date_cell:
            contact_bits.append(f"日付:{date_cell}")
        if flyer:
            contact_bits.append("チラシ送付済")
        if photo:
            contact_bits.append("写真送付済")
        if reply:
            contact_bits.append(f"回答:{reply}")
        note_extra = " / ".join(contact_bits)
        notes_out = f"{notes} | {note_extra}".strip(" |") if note_extra else notes

        lane = infer_property_lane(
            property_area=prop_area,
            city=city,
            notes=notes_out,
            explicit=_cell(r, i_lane),
        )

        row: dict[str, Any] = {
            "id": vid,
            "name": name,
            "area": area or prop_area,
            "prefecture": f"{pref}県" if pref and not pref.endswith("県") else pref,
            "city": city,
            "station": station,
            "property_area": prop_area,
            "property_lane": lane,
            "url": url,
            "contact_url": contact_url,
            "channel": "web_form" if url else "phone",
            "contact_email": email if "@" in email else "",
            "phone": phone,
            "status": status if status in STATUSES else "pending",
            "vacancy_listing_ok": vacancy_ok,
            "precheck_sent_at": "",
            "source": "mgmt_xlsx",
            "notes": notes_out,
            "last_result": result,
            "services": services,
            "discovered_at": now_iso()[:10],
            "contacted_at": "",
            "replied_at": "",
            "alive_checked_at": "",
            "alive_status": "unknown",
            "alive_method": "",
            "alive_note": "",
            "alive_due_days": 180,
        }
        if status == "contacted":
            row["contacted_at"] = date_cell or now_iso()[:10]
            row["precheck_sent_at"] = row["contacted_at"]
        if status == "replied":
            row["replied_at"] = date_cell or now_iso()[:10]
            row["contacted_at"] = date_cell or now_iso()[:10]
            row["precheck_sent_at"] = row["contacted_at"]

        ensure_alive_fields(row, kind="mgmt")
        ensure_precheck_fields(row)

        prev = by_id.get(vid) or by_name.get(name)
        if prev:
            for k in (
                "status",
                "contacted_at",
                "replied_at",
                "precheck_sent_at",
                "vacancy_listing_ok",
                "property_lane",
                "alive_checked_at",
                "alive_status",
                "alive_method",
                "alive_note",
                "last_result",
            ):
                if k in (
                    "alive_checked_at",
                    "alive_status",
                    "alive_method",
                    "alive_note",
                ) and prev.get(k):
                    row[k] = prev[k]
                elif k == "vacancy_listing_ok" and prev.get("vacancy_listing_ok") is not None:
                    if row.get("vacancy_listing_ok") is None:
                        row[k] = prev[k]
                elif k == "property_lane" and prev.get("property_lane") in (
                    "midori_caramel",
                    "both",
                ):
                    # 緑区シード等の手動レーンを維持
                    row[k] = prev[k]
                elif k == "status":
                    # Excel が接触済みなら pending を上書き。手動 skip/invalid は維持
                    if prev.get("status") in ("skip", "invalid") and status == "pending":
                        row["status"] = prev["status"]
                    elif status in ("contacted", "replied", "skip"):
                        row["status"] = status
                    elif prev.get("status") in ("contacted", "replied", "skip", "invalid"):
                        row["status"] = prev["status"]
                elif k in ("contacted_at", "replied_at", "precheck_sent_at"):
                    if status in ("contacted", "replied") and row.get(k):
                        pass  # Excel 日付を優先
                    elif prev.get(k):
                        row[k] = prev[k]
                elif k == "last_result" and prev.get("last_result") and not result:
                    row["last_result"] = prev["last_result"]
            row["id"] = prev["id"]
            if str(prev.get("source") or "").startswith("midori"):
                row["source"] = prev["source"]
            by_id[row["id"]] = row
            updated += 1
        else:
            by_id[vid] = row
            added += 1

    vendors = list(by_id.values())
    vendors.sort(key=lambda v: str(v.get("id") or ""))
    settings = existing.get("settings") or {}
    settings.setdefault("daily_outreach_limit", 2)
    settings.setdefault("alive_due_days", 180)
    settings.setdefault("source_xlsx", str(path))
    out_data = {"settings": settings, "vendors": vendors}
    result = {
        "ok": True,
        "sheet": sheet_name,
        "path": str(path),
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "total": len(vendors),
        "dry_run": dry_run,
    }
    if not dry_run:
        save_list(out_data)
    return result


def summary() -> dict[str, Any]:
    data = load_list()
    vendors = [v for v in (data.get("vendors") or []) if isinstance(v, dict)]
    counts: dict[str, int] = {}
    by_lane: dict[str, int] = {}
    alive_ok_n = 0
    vacancy_ok_n = 0
    for v in vendors:
        ensure_alive_fields(v, kind="mgmt")
        ensure_precheck_fields(v)
        st = str(v.get("status") or "pending")
        counts[st] = counts.get(st, 0) + 1
        lane = str(v.get("property_lane") or "kita_shiga")
        by_lane[lane] = by_lane.get(lane, 0) + 1
        if is_alive_ok(v, kind="mgmt"):
            alive_ok_n += 1
        if v.get("vacancy_listing_ok") is True:
            vacancy_ok_n += 1
    return {
        "ok": True,
        "total": len(vendors),
        "by_status": counts,
        "by_lane": by_lane,
        "alive_ok": alive_ok_n,
        "vacancy_listing_ok": vacancy_ok_n,
        "yaml_exists": LIST_PATH.is_file(),
        "settings": data.get("settings") or {},
        "share_urls": load_share_urls(),
    }


def _lane_match(v: dict[str, Any], lane: str) -> bool:
    pl = str(v.get("property_lane") or "kita_shiga")
    if lane == "both":
        return True
    return pl == lane or pl == "both"


def next_pending(
    *,
    limit: int,
    balanced: bool = False,
    lane: str | None = None,
) -> list[dict[str, Any]]:
    data = load_list()
    pending: list[dict[str, Any]] = []
    for v in data.get("vendors") or []:
        if not isinstance(v, dict):
            continue
        if str(v.get("status") or "pending") not in ("pending", "discovered"):
            continue
        ensure_alive_fields(v, kind="mgmt")
        ensure_precheck_fields(v)
        if lane and not _lane_match(v, lane):
            continue
        pending.append(v)

    share = load_share_urls()

    def enrich(v: dict[str, Any]) -> dict[str, Any]:
        pl = str(v.get("property_lane") or "kita_shiga")
        prop_url = share.get(pl if pl != "both" else "kita_shiga") or ""
        if pl == "both":
            prop_url = share.get("kita_shiga") or share.get("midori_caramel") or ""
        return {
            **v,
            "precheck": {
                "area_phrase": AREA_PHRASE.get(pl, AREA_PHRASE["kita_shiga"]),
                "property_folder_url": prop_url,
                "common_folder_url": share.get("common") or "",
                "subject": (
                    "【空室対策のご相談】名古屋市緑区にアパートを所有している松野です"
                    if pl == "midori_caramel"
                    else "【空室対策のご相談】名古屋市北区にアパートを所有している松野です"
                ),
            },
        }

    if not balanced or limit < 2 or lane:
        return [enrich(v) for v in pending[:limit]]

    kita_ex = [v for v in pending if v.get("property_lane") == "kita_shiga"]
    midori_ex = [v for v in pending if v.get("property_lane") == "midori_caramel"]
    both = [v for v in pending if v.get("property_lane") == "both"]
    out: list[dict[str, Any]] = []
    used: set[str] = set()
    for pool in (kita_ex, midori_ex):
        for v in pool:
            vid = str(v.get("id"))
            if vid in used:
                continue
            out.append(enrich(v))
            used.add(vid)
            break
        if len(out) >= limit:
            break
    for v in both + kita_ex + midori_ex + pending:
        if len(out) >= limit:
            break
        vid = str(v.get("id"))
        if vid in used:
            continue
        out.append(enrich(v))
        used.add(vid)
    return out


def vacancy_eligible(*, lane: str | None = None) -> list[dict[str, Any]]:
    data = load_list()
    out: list[dict[str, Any]] = []
    for v in data.get("vendors") or []:
        if not isinstance(v, dict):
            continue
        ensure_precheck_fields(v)
        if v.get("vacancy_listing_ok") is not True:
            continue
        if lane and not _lane_match(v, lane):
            continue
        out.append(v)
    return out


def mark_vendor(
    vid: str,
    *,
    status: str,
    note: str = "",
    result: str = "",
    vacancy_listing_ok: str | None = None,
    kodate_mgmt_ok: str | None = None,
    property_lane: str | None = None,
    dry_run: bool,
) -> dict[str, Any]:
    if status not in STATUSES:
        return {"ok": False, "error": f"invalid status: {status}"}
    data = load_list()
    by_id = vendor_index(data)
    v = by_id.get(vid)
    if not v:
        return {"ok": False, "error": f"vendor not found: {vid}"}
    today = now_iso()[:10]
    v["status"] = status
    if note:
        v["notes"] = note if not v.get("notes") else f"{v['notes']} | {note}"
    if result:
        v["last_result"] = result
    if status == "contacted":
        v["contacted_at"] = today
        v["precheck_sent_at"] = today
    if status == "replied":
        v["replied_at"] = today
        v["contacted_at"] = v.get("contacted_at") or today
    if vacancy_listing_ok is not None and vacancy_listing_ok != "":
        s = str(vacancy_listing_ok).strip().lower()
        if s in ("true", "1", "ok", "yes", "可"):
            v["vacancy_listing_ok"] = True
        elif s in ("false", "0", "ng", "no", "不可"):
            v["vacancy_listing_ok"] = False
            if status == "replied":
                v["status"] = "skip"
        elif s in ("null", "none", "-"):
            v["vacancy_listing_ok"] = None
    if kodate_mgmt_ok is not None and kodate_mgmt_ok != "":
        services = v.get("services") if isinstance(v.get("services"), dict) else {}
        s = str(kodate_mgmt_ok).strip().lower()
        services["kodate_mgmt_ok"] = s in ("true", "1", "ok", "yes", "可")
        v["services"] = services
    if property_lane and property_lane in LANES:
        v["property_lane"] = property_lane
    v["updated_at"] = now_iso()
    ensure_alive_fields(v, kind="mgmt")
    ensure_precheck_fields(v)
    if not dry_run:
        save_list(data)
    return {"ok": True, "vendor": v, "dry_run": dry_run}


def mark_vendor_alive(
    vid: str,
    *,
    alive_status: str,
    method: str = "phone",
    note: str = "",
    dry_run: bool,
) -> dict[str, Any]:
    st = (alive_status or "").strip().lower()
    if st not in ("ok", "fail", "unknown"):
        return {"ok": False, "error": f"invalid alive_status: {alive_status}"}
    data = load_list()
    by_id = vendor_index(data)
    v = by_id.get(vid)
    if not v:
        return {"ok": False, "error": f"vendor not found: {vid}"}
    mark_alive_fields(v, status=st, method=method, note=note, kind="mgmt")
    if not dry_run:
        save_list(data)
    return {
        "ok": True,
        "vendor": v,
        "alive_effective": effective_alive_status(v, kind="mgmt"),
        "dry_run": dry_run,
    }


def merge_append(vendors: list[Any], *, dry_run: bool) -> dict[str, Any]:
    data = load_list()
    by_id = vendor_index(data)
    added = 0
    for raw in vendors:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or "").strip()
        if not name:
            continue
        vid = str(raw.get("id") or "").strip() or slug_id(
            name, str(raw.get("area") or "")
        )
        if vid in by_id:
            continue
        row = {
            "id": vid,
            "name": name,
            "area": raw.get("area") or "",
            "prefecture": raw.get("prefecture") or "",
            "city": raw.get("city") or "",
            "station": raw.get("station") or "",
            "property_area": raw.get("property_area") or "",
            "property_lane": raw.get("property_lane")
            or infer_property_lane(
                property_area=str(raw.get("property_area") or ""),
                city=str(raw.get("city") or ""),
                notes=str(raw.get("notes") or ""),
            ),
            "url": raw.get("url") or "",
            "contact_url": raw.get("contact_url") or raw.get("url") or "",
            "channel": raw.get("channel") or "web_form",
            "contact_email": raw.get("contact_email") or "",
            "phone": raw.get("phone") or "",
            "status": raw.get("status") or "discovered",
            "vacancy_listing_ok": raw.get("vacancy_listing_ok", None),
            "precheck_sent_at": "",
            "source": raw.get("source") or "grok_discovery",
            "notes": raw.get("notes") or "",
            "services": raw.get("services") or {},
            "discovered_at": raw.get("discovered_at") or now_iso()[:10],
            "contacted_at": "",
            "replied_at": "",
            "last_result": "",
            "alive_due_days": 180,
        }
        ensure_alive_fields(row, kind="mgmt")
        ensure_precheck_fields(row)
        by_id[vid] = row
        added += 1
    data["vendors"] = list(by_id.values())
    if not dry_run:
        save_list(data)
    return {"ok": True, "added": added, "total": len(data["vendors"]), "dry_run": dry_run}


def backfill_lanes(*, dry_run: bool) -> dict[str, Any]:
    """既存 YAML に property_lane / vacancy_listing_ok を埋める。"""
    data = load_list()
    n = 0
    for v in data.get("vendors") or []:
        if not isinstance(v, dict):
            continue
        before = v.get("property_lane")
        ensure_precheck_fields(v)
        if v.get("property_lane") != before or "vacancy_listing_ok" not in v:
            n += 1
        # 緑区補強: 名前にキャラメル関連は midori（現状リストは北区中心）
        notes = str(v.get("notes") or "")
        if "緑区" in notes or "キャラメル" in notes:
            v["property_lane"] = "midori_caramel"
    if not dry_run:
        save_list(data)
    return {"ok": True, "touched": n, "total": len(data.get("vendors") or []), "dry_run": dry_run}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--summary", action="store_true")
    ap.add_argument("--import-xlsx", metavar="PATH", nargs="?", const=str(DEFAULT_XLSX))
    ap.add_argument("--merge", action="store_true", default=True)
    ap.add_argument("--no-merge", action="store_true")
    ap.add_argument("--next", type=int, default=0)
    ap.add_argument("--balanced", action="store_true", help="北区1+緑区1 を優先")
    ap.add_argument("--lane", default="", help="kita_shiga|midori_caramel|both")
    ap.add_argument("--vacancy-eligible", action="store_true")
    ap.add_argument("--backfill-lanes", action="store_true")
    ap.add_argument("--mark", metavar="ID")
    ap.add_argument("--status", default="contacted")
    ap.add_argument("--note", default="")
    ap.add_argument("--result", default="")
    ap.add_argument("--vacancy-listing-ok", default="")
    ap.add_argument("--kodate-mgmt-ok", default="")
    ap.add_argument("--property-lane", default="")
    ap.add_argument("--mark-alive", metavar="ID")
    ap.add_argument("--alive-status", default="ok")
    ap.add_argument("--alive-method", default="phone")
    ap.add_argument("--alive-queue", action="store_true")
    ap.add_argument("--limit", type=int, default=2)
    ap.add_argument("--merge-append", metavar="PATH")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    merge = not args.no_merge

    if args.backfill_lanes:
        out = backfill_lanes(dry_run=args.dry_run)
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 0

    if args.vacancy_eligible:
        items = vacancy_eligible(lane=args.lane or None)
        print(
            json.dumps(
                {"ok": True, "count": len(items), "lane": args.lane or "all", "vendors": items},
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if args.alive_queue:
        data = load_list()
        vendors = [v for v in (data.get("vendors") or []) if isinstance(v, dict)]
        items = build_alive_queue(vendors, kind="mgmt", limit=max(1, args.limit))
        print(
            json.dumps(
                {
                    "ok": True,
                    "kind": "mgmt",
                    "count": len(items),
                    "vendors": [
                        {
                            "id": v.get("id"),
                            "name": v.get("name"),
                            "phone": v.get("phone"),
                            "url": v.get("url"),
                            "property_lane": v.get("property_lane"),
                            "alive_status": v.get("alive_status"),
                            "alive_effective": effective_alive_status(v, kind="mgmt"),
                        }
                        for v in items
                    ],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if args.mark_alive:
        out = mark_vendor_alive(
            args.mark_alive,
            alive_status=args.alive_status,
            method=args.alive_method,
            note=args.note,
            dry_run=args.dry_run,
        )
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 0 if out.get("ok") else 1

    if args.import_xlsx:
        out = import_mgmt_xlsx(
            Path(args.import_xlsx), dry_run=args.dry_run, merge=merge
        )
        print(json.dumps(out, ensure_ascii=False, indent=2))
        print(f"MGMT_VENDOR_LIST_RESULT:{json.dumps(out, ensure_ascii=False)}")
        return 0

    if args.merge_append:
        raw = yaml.safe_load(Path(args.merge_append).read_text(encoding="utf-8")) or {}
        vendors = raw.get("vendors") if isinstance(raw, dict) else raw
        if not isinstance(vendors, list):
            print(json.dumps({"ok": False, "error": "no vendors"}, ensure_ascii=False))
            return 1
        out = merge_append(vendors, dry_run=args.dry_run)
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 0

    if args.mark:
        out = mark_vendor(
            args.mark,
            status=args.status,
            note=args.note,
            result=args.result,
            vacancy_listing_ok=args.vacancy_listing_ok or None,
            kodate_mgmt_ok=args.kodate_mgmt_ok or None,
            property_lane=args.property_lane or None,
            dry_run=args.dry_run,
        )
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 0 if out.get("ok") else 1

    if args.next > 0:
        items = next_pending(
            limit=args.next,
            balanced=bool(args.balanced),
            lane=args.lane or None,
        )
        print(
            json.dumps(
                {
                    "ok": True,
                    "count": len(items),
                    "balanced": bool(args.balanced),
                    "lane": args.lane or None,
                    "vendors": items,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    out = summary()
    print(json.dumps(out, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
