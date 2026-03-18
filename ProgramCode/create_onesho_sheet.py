# -*- coding: utf-8 -*-
"""おねしょチェック表をExcelで作成（A4縦・1シートに2ヶ月・週ごと折り返し・たまき・さわ）"""
import calendar
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# A4印刷で見やすいサイズ（幅・高さをやや大きめに）
COL_LABEL_WIDTH = 7.0   # A列（日付/たまき/さわ）
COL_DAY_WIDTH = 5.5     # B〜H（曜日セル）
ROW_TITLE_HEIGHT = 20
ROW_WEEKDAY_HEIGHT = 18
ROW_DATE_HEIGHT = 18
ROW_NAME_HEIGHT = 20
FONT_TITLE_SIZE = 12
FONT_CELL_SIZE = 10

weekday_names = ["月", "火", "水", "木", "金", "土", "日"]
thin = Side(style="thin", color="000000")


def add_month_block(ws, year: int, month: int, start_row: int, is_first: bool):
    """
    指定シートに1ヶ月分のブロックを追加する。
    is_first=True のときはタイトル行を描く。戻り値は次の開始行。
    """
    cal = calendar.Calendar(calendar.MONDAY)
    weeks = list(cal.monthdays2calendar(year, month))
    max_weeks = 5
    while len(weeks) < max_weeks:
        weeks.append([(0, i) for i in range(7)])

    row = start_row

    if is_first:
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value=f"{year}年{month}月")
        ws.cell(row=row, column=1).font = Font(bold=True, size=FONT_TITLE_SIZE)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = ROW_TITLE_HEIGHT
        row += 1

    for w in weeks[:max_weeks]:
        # 曜日行
        ws.cell(row=row, column=1, value="").border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for d in range(7):
            c = ws.cell(row=row, column=2 + d, value=weekday_names[d])
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True, size=FONT_CELL_SIZE)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = ROW_WEEKDAY_HEIGHT
        row += 1

        # 日付行
        ws.cell(row=row, column=1, value="日付").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).font = Font(bold=True, size=FONT_CELL_SIZE)
        ws.cell(row=row, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for d in range(7):
            day = w[d][0]
            c = ws.cell(row=row, column=2 + d, value=day if day else "")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=FONT_CELL_SIZE)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = ROW_DATE_HEIGHT
        row += 1

        # たまき行
        ws.cell(row=row, column=1, value="たまき").font = Font(bold=True, size=FONT_CELL_SIZE)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for d in range(7):
            c = ws.cell(row=row, column=2 + d)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = ROW_NAME_HEIGHT
        row += 1

        # さわ行
        ws.cell(row=row, column=1, value="さわ").font = Font(bold=True, size=FONT_CELL_SIZE)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for d in range(7):
            c = ws.cell(row=row, column=2 + d)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = ROW_NAME_HEIGHT
        row += 1

    return row


def create_two_month_sheet(wb, year: int, month1: int, month2: int, sheet_title: str):
    """1シートに2ヶ月を配置し、月の間に1行空ける。"""
    ws = wb.create_sheet(title=sheet_title, index=0)

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    row = 1
    row = add_month_block(ws, year, month1, row, is_first=True)
    # 4月と5月（または6月と7月）の間を1行あける
    row += 1
    row = add_month_block(ws, year, month2, row, is_first=True)  # 2ヶ月目もタイトル行を描く

    ws.column_dimensions["A"].width = COL_LABEL_WIDTH
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = COL_DAY_WIDTH

    return ws


def main():
    out_path = Path(__file__).parent / "おねしょチェック表.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    year = date.today().year

    # シート1: 4月・5月（間に1行空け）
    create_two_month_sheet(wb, year, 4, 5, "4月・5月")
    # シート2: 6月・7月（間に1行空け）
    create_two_month_sheet(wb, year, 6, 7, "6月・7月")

    wb.save(out_path)
    print(f"保存しました: {out_path}")


if __name__ == "__main__":
    main()
