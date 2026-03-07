#!/usr/bin/env python3
"""
Excel to Markdown 変換スクリプト
Excelファイル(.xlsx)からデータを読み取り、Markdownテーブル形式で保存する。
全シートを自動的に変換し、シートごとに見出しで区切る。

使い方:
    python3 excel_to_md.py input.xlsx
    python3 excel_to_md.py input.xlsx -o ~/Downloads/
"""

import argparse
import os
import sys
from datetime import datetime, date, time

import openpyxl


def cell_to_str(value) -> str:
    """セルの値を文字列に変換する"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float):
        # 整数値の場合は小数点を省く
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def sheet_to_markdown(sheet) -> str:
    """ワークシートのデータをMarkdownテーブル形式に変換する"""
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return "*(このシートにはデータがありません)*"

    # 全行を文字列に変換
    str_rows = []
    for row in rows:
        str_rows.append([cell_to_str(cell) for cell in row])

    # 最大列数を取得
    max_cols = max(len(row) for row in str_rows)
    if max_cols == 0:
        return "*(このシートにはデータがありません)*"

    # 列数を揃える
    for row in str_rows:
        while len(row) < max_cols:
            row.append("")

    # 先頭の完全空行をスキップ
    start_idx = 0
    for i, row in enumerate(str_rows):
        if any(cell.strip() for cell in row):
            start_idx = i
            break

    # 末尾の完全空行をスキップ
    end_idx = len(str_rows)
    for i in range(len(str_rows) - 1, -1, -1):
        if any(cell.strip() for cell in str_rows[i]):
            end_idx = i + 1
            break

    str_rows = str_rows[start_idx:end_idx]

    if not str_rows:
        return "*(このシートにはデータがありません)*"

    # Markdownテーブルを構築（パイプ内のパイプ文字をエスケープ）
    lines = []

    # ヘッダー行（1行目）
    header = [cell.replace("|", "\\|") for cell in str_rows[0]]
    lines.append("| " + " | ".join(header) + " |")

    # 区切り行
    lines.append("| " + " | ".join(["---"] * max_cols) + " |")

    # データ行
    for row in str_rows[1:]:
        escaped = [cell.replace("|", "\\|") for cell in row]
        lines.append("| " + " | ".join(escaped) + " |")

    return "\n".join(lines)


def excel_to_markdown(excel_path: str, output_path: str) -> str:
    """ExcelファイルをMarkdownに変換して保存する"""
    if not os.path.exists(excel_path):
        print(f"エラー: ファイルが見つかりません: {excel_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    md_parts = []

    # ファイル名をタイトルとして追加
    excel_name = os.path.splitext(os.path.basename(excel_path))[0]
    md_parts.append(f"# {excel_name}\n")

    print(f"処理中: {excel_path} ({len(wb.sheetnames)}シート)")

    for name in wb.sheetnames:
        sheet = wb[name]
        md_parts.append(f"## {name}\n")
        md_table = sheet_to_markdown(sheet)
        md_parts.append(md_table)
        md_parts.append("")  # 空行で区切り
        print(f"  変換済み: {name}")

    wb.close()

    md_content = "\n".join(md_parts)

    # ファイルに書き出し
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"変換完了: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Excelファイル(.xlsx)をMarkdown(.md)に変換するツール"
    )
    parser.add_argument(
        "excel_path",
        help="変換するExcelファイルのパス",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        help="出力先ディレクトリ（省略時はExcelと同じディレクトリ）",
        default=None,
    )

    args = parser.parse_args()

    # Excelパスを絶対パスに変換
    excel_path = os.path.abspath(os.path.expanduser(args.excel_path))

    if not os.path.exists(excel_path):
        print(f"エラー: ファイルが見つかりません: {excel_path}", file=sys.stderr)
        sys.exit(1)

    # 出力パスを決定
    excel_basename = os.path.splitext(os.path.basename(excel_path))[0]
    md_filename = f"{excel_basename}.md"

    if args.output_dir:
        output_dir = os.path.abspath(os.path.expanduser(args.output_dir))
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, md_filename)
    else:
        output_path = os.path.join(os.path.dirname(excel_path), md_filename)

    excel_to_markdown(excel_path, output_path)


if __name__ == "__main__":
    main()
